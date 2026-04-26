"""Read DDE-produced xlsx workbooks into a stream of dicts.

DDE's writer (``requirements_extractor/writer.py``) emits the column
order::

    # | ID | Source File | Heading Trail | Section / Topic | Row Ref |
    Block Ref | Primary Actor | Secondary Actors | Requirement | Type |
    Polarity | Keywords | Confidence | Notes | Context

This module loads that shape and yields one ``dict`` per requirement
row. Consumer tools (compliance-matrix, nimbus-skeleton) wrap each
dict in their own ``DDERow`` dataclass — keeping the consumer-side
type narrow to just the fields each tool actually uses.

The loader matches columns by **header name** (case-insensitive,
whitespace-collapsed) rather than position so a future DDE schema
change that adds or reorders columns doesn't break consumers.

Helper helpers (added 2026-04-25):

* ``load_into(path, row_factory, fields=None)`` — common loader
  pattern: iterate, filter to a field subset, hand each dict to a
  factory. Both consumer tools have nearly-identical loops; this
  centralises them.
* ``find_sidecar(input_path, *, suffix)`` — convention helper for
  "look up the actors xlsx beside the input file." Returns the
  expected sidecar path if it exists, else ``None``.
"""

from __future__ import annotations

from pathlib import Path
from typing import Callable, Dict, Iterable, Iterator, List, Optional, TypeVar

from openpyxl import load_workbook


T = TypeVar("T")


# Header name (normalised) → canonical attribute name. Adding a new
# column to DDE? Add the (lowercase) header → attr mapping here and
# both consumer tools see it for free.
HEADER_MAP: Dict[str, str] = {
    "id": "stable_id",
    "stable id": "stable_id",
    "source file": "source_file",
    "heading trail": "heading_trail",
    "section / topic": "section",
    "section/topic": "section",
    "section": "section",
    "row ref": "row_ref",
    "block ref": "block_ref",
    "primary actor": "primary_actor",
    "secondary actors": "secondary_actors",
    "requirement": "text",
    "clause": "text",
    "type": "req_type",
    "polarity": "polarity",
    "keywords": "keywords",
    "confidence": "confidence",
    "notes": "notes",
    "context": "context",
}


def normalise_header(value: object) -> str:
    """Lowercase + whitespace-collapse a header cell. Returns ``""`` for
    ``None`` / empty input so the caller can compare safely."""

    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def iter_dde_records(path: str | Path) -> Iterator[Dict[str, str]]:
    """Yield one dict per data row of a DDE xlsx workbook.

    Each dict has whichever subset of ``HEADER_MAP`` values the workbook
    actually populates (the loader skips ``None`` cells). Rows missing
    both ``stable_id`` and ``text`` are silently skipped — these are
    typically summary footers DDE appends.

    Raises ``ValueError`` if the workbook lacks both an ``ID`` column
    and a ``Requirement``/``Clause`` column — without those there's no
    way to identify or quote the rows.
    """

    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return

        col_map: Dict[int, str] = {}
        for idx, header in enumerate(header_row):
            attr = HEADER_MAP.get(normalise_header(header))
            if attr is not None:
                col_map[idx] = attr

        attrs = set(col_map.values())
        if "stable_id" not in attrs or "text" not in attrs:
            raise ValueError(
                f"{path.name} is missing required columns (need both "
                "'ID' and 'Requirement'/'Clause')"
            )

        for raw in rows_iter:
            record: Dict[str, str] = {}
            for idx, attr in col_map.items():
                if idx >= len(raw):
                    continue
                value = raw[idx]
                if value is None:
                    continue
                record[attr] = str(value).strip()
            if not record.get("stable_id") or not record.get("text"):
                continue
            yield record
    finally:
        wb.close()


def load_dde_records(path: str | Path) -> List[Dict[str, str]]:
    """List-form convenience wrapper around ``iter_dde_records``."""

    return list(iter_dde_records(path))


def iter_actor_records(path: str | Path) -> Iterator[Dict[str, list[str]]]:
    """Yield ``{actor: str, aliases: [str]}`` dicts from a DDE actors xlsx.

    Returns an empty iterator (no yields) if the workbook isn't
    actors-shaped — the consumer treats a missing actors table as
    'derive everything from the requirements side', so silently
    degrading is the right behaviour rather than raising.
    """

    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return

        headers = [normalise_header(h) for h in header_row]
        actor_idx: Optional[int] = next(
            (i for i, h in enumerate(headers) if h == "actor"), None
        )
        aliases_idx: Optional[int] = next(
            (i for i, h in enumerate(headers) if h == "aliases"), None
        )

        if actor_idx is None:
            return

        for raw in rows_iter:
            if actor_idx >= len(raw):
                continue
            actor_value = raw[actor_idx]
            if not actor_value:
                continue
            actor = str(actor_value).strip()
            aliases: List[str] = []
            if (
                aliases_idx is not None
                and aliases_idx < len(raw)
                and raw[aliases_idx]
            ):
                aliases = [
                    a.strip()
                    for a in str(raw[aliases_idx]).split(",")
                    if a.strip()
                ]
            yield {"actor": actor, "aliases": aliases}
    finally:
        wb.close()


def load_actor_aliases(path: str | Path) -> Dict[str, List[str]]:
    """Return ``{canonical: [aliases]}`` from a DDE actors xlsx."""

    return {rec["actor"]: rec["aliases"] for rec in iter_actor_records(path)}


def load_into(
    path: str | Path,
    row_factory: Callable[..., T],
    *,
    fields: Optional[Iterable[str]] = None,
) -> List[T]:
    """Load a DDE xlsx and project each row through ``row_factory``.

    Both downstream tools (compliance-matrix, nimbus-skeleton) had
    near-identical "iterate, filter to known fields, build domain row"
    loops. This helper takes a callable — typically a dataclass
    constructor — and a whitelist of field names. Each yielded dict
    from ``iter_dde_records`` is filtered to the whitelisted keys
    (if any) and passed as keyword arguments to ``row_factory``.

    Pass ``fields=None`` to forward every key in the dict (no
    filtering — useful when the row type accepts arbitrary kwargs).
    """

    if fields is None:
        return [row_factory(**rec) for rec in iter_dde_records(path)]

    allowed = set(fields)
    return [
        row_factory(**{k: v for k, v in rec.items() if k in allowed})
        for rec in iter_dde_records(path)
    ]


def find_sidecar(
    input_path: str | Path,
    *,
    suffix: str,
    extension: str = ".xlsx",
) -> Optional[Path]:
    """Look for a sidecar file next to ``input_path`` and return it if present.

    Convention: a sidecar is a file in the same directory as
    ``input_path``, with the same stem plus a discriminating suffix
    and an extension. For example, given ``contract.xlsx`` and
    ``suffix="_actors"``, this looks for ``contract_actors.xlsx``.

    Returns the ``Path`` if the sidecar exists, otherwise ``None``.
    Both tools today use this convention for the actors workbook;
    centralising the lookup means a typo in one consumer doesn't
    silently miss a real sidecar.
    """

    input_path = Path(input_path)
    candidate = input_path.with_name(f"{input_path.stem}{suffix}{extension}")
    return candidate if candidate.is_file() else None
