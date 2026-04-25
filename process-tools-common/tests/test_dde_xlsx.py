"""Tests for the DDE xlsx loader.

Covers the header-name matching (case / whitespace tolerance), the
required-column check, and the silent skip of non-DDE-shaped actors
workbooks.
"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from process_tools_common.dde_xlsx import (
    HEADER_MAP,
    iter_actor_records,
    iter_dde_records,
    load_actor_aliases,
    load_dde_records,
    normalise_header,
)


def _write_minimal_dde_xlsx(path: Path, headers, rows):
    wb = Workbook()
    ws = wb.active
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    wb.save(path)


class TestNormaliseHeader(unittest.TestCase):
    def test_collapses_whitespace_and_lowercases(self) -> None:
        self.assertEqual(normalise_header("  Source File  "), "source file")
        self.assertEqual(normalise_header("Source\tFile"), "source file")
        self.assertEqual(normalise_header("SECTION / TOPIC"), "section / topic")

    def test_handles_none(self) -> None:
        self.assertEqual(normalise_header(None), "")


class TestIterDdeRecords(unittest.TestCase):
    def test_round_trip(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            xlsx = Path(td) / "test.xlsx"
            _write_minimal_dde_xlsx(
                xlsx,
                headers=["#", "ID", "Source File", "Primary Actor", "Requirement", "Polarity"],
                rows=[
                    [1, "REQ-AAA1", "spec.docx", "Operator", "Operator shall log in.", "Positive"],
                    [2, "REQ-AAA2", "spec.docx", "System", "System shall record event.", "Positive"],
                ],
            )
            records = load_dde_records(xlsx)
            self.assertEqual(len(records), 2)
            self.assertEqual(records[0]["stable_id"], "REQ-AAA1")
            self.assertEqual(records[0]["text"], "Operator shall log in.")
            self.assertEqual(records[0]["primary_actor"], "Operator")
            self.assertEqual(records[1]["polarity"], "Positive")

    def test_skips_empty_rows(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            xlsx = Path(td) / "test.xlsx"
            _write_minimal_dde_xlsx(
                xlsx,
                headers=["ID", "Requirement"],
                rows=[
                    ["REQ-1", "First."],
                    [None, None],
                    ["REQ-2", "Second."],
                ],
            )
            records = load_dde_records(xlsx)
            self.assertEqual(len(records), 2)

    def test_missing_columns_raises(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            xlsx = Path(td) / "bad.xlsx"
            _write_minimal_dde_xlsx(
                xlsx,
                headers=["Foo", "Bar"],
                rows=[["a", "b"]],
            )
            with self.assertRaises(ValueError):
                load_dde_records(xlsx)

    def test_header_matching_is_case_insensitive(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            xlsx = Path(td) / "messy.xlsx"
            _write_minimal_dde_xlsx(
                xlsx,
                headers=["  ID  ", "REQUIREMENT", "polarity"],
                rows=[["REQ-X", "X.", "Negative"]],
            )
            records = load_dde_records(xlsx)
            self.assertEqual(records[0]["polarity"], "Negative")


class TestIterActorRecords(unittest.TestCase):
    def test_basic_actors(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            xlsx = Path(td) / "actors.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="Actor")
            ws.cell(row=1, column=2, value="Aliases")
            ws.cell(row=2, column=1, value="Operator")
            ws.cell(row=2, column=2, value="op, the operator")
            ws.cell(row=3, column=1, value="System")
            ws.cell(row=3, column=2, value="")
            wb.save(xlsx)

            aliases = load_actor_aliases(xlsx)
            self.assertEqual(aliases["Operator"], ["op", "the operator"])
            self.assertEqual(aliases["System"], [])

    def test_non_actors_workbook_returns_empty(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            xlsx = Path(td) / "not_actors.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="Foo")
            ws.cell(row=2, column=1, value="bar")
            wb.save(xlsx)

            self.assertEqual(load_actor_aliases(xlsx), {})


class TestHeaderMap(unittest.TestCase):
    def test_required_attrs_present(self) -> None:
        attrs = set(HEADER_MAP.values())
        for required in ("stable_id", "text", "primary_actor", "polarity"):
            self.assertIn(required, attrs)


if __name__ == "__main__":
    unittest.main()
