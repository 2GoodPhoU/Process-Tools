"""Write extracted requirements to a formatted .xlsx file."""

from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Requirement

COLUMNS = [
    ("#", 6),
    ("ID", 14),
    ("Source File", 28),
    ("Heading Trail", 30),
    ("Section / Topic", 22),
    ("Row Ref", 18),
    ("Block Ref", 18),
    ("Primary Actor", 22),
    ("Secondary Actors", 26),
    ("Requirement", 70),
    ("Type", 8),
    ("Polarity", 10),
    ("Keywords", 20),
    ("Confidence", 12),
    ("Notes", 36),
]

HEADER_FILL = PatternFill("solid", start_color="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
SOFT_FILL = PatternFill("solid", start_color="FFF2CC")  # light yellow
# Negative requirements (shall-not etc.) get a tinted background so they
# stand out during review.  Subtle so the Soft highlight still dominates.
NEGATIVE_FILL = PatternFill("solid", start_color="F4CCCC")  # light red


def write_requirements(
    requirements: Sequence[Requirement],
    output_path: Path,
    *,
    sheet_name: str = "Requirements",
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Body rows
    for row_idx, req in enumerate(requirements, start=2):
        values = [
            req.order,
            req.stable_id,
            req.source_file,
            req.heading_trail,
            req.section_topic,
            req.row_ref,
            req.block_ref,
            req.primary_actor,
            req.secondary_actors_str,
            req.text,
            req.req_type,
            req.polarity,
            req.keywords_str,
            req.confidence,
            req.notes,
        ]
        # Pick a row fill: Negative polarity beats Soft, since flagging
        # "shall not / must not" language for a reviewer matters more than
        # the advisory-vs-binding distinction.
        row_fill = None
        if req.polarity == "Negative":
            row_fill = NEGATIVE_FILL
        elif req.req_type == "Soft":
            row_fill = SOFT_FILL
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_fill is not None:
                cell.fill = row_fill

    # Autofilter across the populated range
    if requirements:
        last_col = get_column_letter(len(COLUMNS))
        last_row = len(requirements) + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Add a summary sheet
    summary = wb.create_sheet("Summary")
    _write_summary(summary, requirements)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


def _write_summary(ws, requirements: Sequence[Requirement]) -> None:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14

    rows = [
        ("Total requirements", len(requirements)),
        ("Hard requirements", sum(1 for r in requirements if r.req_type == "Hard")),
        ("Soft (needs review)", sum(1 for r in requirements if r.req_type == "Soft")),
        ("", ""),
        ("Top actors", ""),
    ]
    for label, value in rows:
        ws.append([label, value])

    # Count by primary actor
    counts: dict[str, int] = {}
    for r in requirements:
        key = r.primary_actor or "(no primary actor)"
        counts[key] = counts.get(key, 0) + 1
    for actor, count in sorted(counts.items(), key=lambda kv: -kv[1]):
        ws.append([actor, count])

    # Bold the title rows
    bold = Font(name="Arial", bold=True)
    for cell in ws["A"]:
        cell.font = bold
