"""Pure-Python helpers used by the Tk GUI.

The Tk UI itself is hard to drive in headless CI, so everything that
*doesn't* need a live Tk root lives here and gets unit-tested directly.
This includes:

* ``GuiSettings`` — a dataclass + JSON round-trip for persisting window
  size and last-used paths between runs.
* ``dedupe_paths`` — a resolved-path aware dedup that treats
  ``./a/../a/spec.docx`` and ``a/spec.docx`` as the same input.
* ``write_actors_template`` — generate an example actors .xlsx so users
  don't have to hand-roll one.

The Tk surface in ``gui.py`` imports and composes these.
"""

from __future__ import annotations

import json
from dataclasses import MISSING, asdict, dataclass, field, fields
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional


# ---------------------------------------------------------------------------
# Persistent settings
# ---------------------------------------------------------------------------


#: Extraction modes the GUI knows how to run.  Mirrors the CLI.
_VALID_MODES = frozenset({"requirements", "actors"})


def default_settings_path() -> Path:
    """Return the per-user settings file location.

    Uses the platform convention the *Python stdlib* already picks:
    ``Path.home()`` on every OS.  We don't chase XDG or AppData here
    because the tool's users span Windows, macOS, and Linux and a single
    predictable location is friendlier for hand-inspection.
    """
    return Path.home() / ".requirements_extractor" / "settings.json"


@dataclass
class GuiSettings:
    """Everything the GUI remembers across restarts.

    All fields have safe defaults so a brand-new install (or a corrupted
    settings file) still launches cleanly.  The JSON file is versioned
    via ``schema_version`` so future schema changes can migrate old
    files in place without wiping user preferences.
    """

    schema_version: int = 1

    # Window geometry — stored as a string like "760x560+120+80" that
    # Tk's ``root.geometry()`` consumes directly.
    window_geometry: str = "760x560"

    # Last-used paths (empty string = no preference).
    last_actors_path: str = ""
    last_config_path: str = ""
    last_output_path: str = ""
    last_statement_set_path: str = ""
    last_input_dir: str = ""  # initialdir for "Add file(s)…" / "Add folder…"

    # Checkbox state.
    use_nlp: bool = False
    export_statement_set: bool = False
    open_output_on_done: bool = True  # default to opening — it's the common case
    # Dry-run mirrors the CLI's ``--dry-run`` flag: parse + detect + assign
    # stable IDs, but skip writing any files.  Persisted off by default so
    # a forgetful user doesn't silently disable output across restarts.
    dry_run: bool = False

    # Extraction mode the user last chose in the GUI.  Mirrors the CLI
    # subcommand names so the two surfaces stay in lockstep.  Unknown
    # values coming off disk collapse to "requirements".
    mode: str = "requirements"

    # Recently used docx inputs, most-recent-first.  Capped in _trim_recent.
    recent_inputs: List[str] = field(default_factory=list)

    # --- JSON round-trip ------------------------------------------------- #

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)

    @classmethod
    def from_dict(cls, raw: Dict[str, Any]) -> "GuiSettings":
        """Build from a (possibly partial / malformed) dict.

        Unknown keys are silently ignored so a new release that *drops*
        a field doesn't crash on an old user's settings file.  Missing
        keys fall back to the dataclass default.  Type mismatches (e.g.
        a string where a bool is expected) also fall back, rather than
        throwing — the goal is 'always launch cleanly'.
        """
        kwargs: Dict[str, Any] = {}
        known = {f.name: f for f in fields(cls)}
        for name, f in known.items():
            if name not in raw:
                continue
            value = raw[name]
            # Dataclass field annotations are strings under
            # ``from __future__ import annotations``, so type-check by
            # comparing against the default's concrete runtime type.
            if f.default is not MISSING:
                default = f.default
            elif f.default_factory is not MISSING:  # type: ignore[misc]
                default = f.default_factory()  # type: ignore[misc]
            else:  # required field — no defaults exist for any of ours, but be safe
                default = None
            # Order matters: ``bool`` is a subclass of ``int`` in Python,
            # so the bool guard must run before the int guard.
            if isinstance(default, bool):
                if not isinstance(value, bool):
                    continue
            elif isinstance(default, int):
                if not isinstance(value, int) or isinstance(value, bool):
                    continue
            elif isinstance(default, str):
                if not isinstance(value, str):
                    continue
            elif isinstance(default, list):
                if not isinstance(value, list):
                    continue
            kwargs[name] = value
        inst = cls(**kwargs)
        # Guard: mode must be one of the CLI subcommand names.  A value
        # left over from a future release (or hand-edited) shouldn't
        # wedge the GUI — fall back to the default instead.
        if inst.mode not in _VALID_MODES:
            inst.mode = "requirements"
        return inst

    # --- Persistence ----------------------------------------------------- #

    @classmethod
    def load(cls, path: Optional[Path] = None) -> "GuiSettings":
        """Load settings from ``path`` (defaults to ``default_settings_path``).

        Never raises: a missing file, unreadable file, or malformed JSON
        all collapse into 'return defaults'.  The GUI must always start.
        """
        p = path or default_settings_path()
        try:
            raw = json.loads(p.read_text(encoding="utf-8"))
        except FileNotFoundError:
            return cls()
        except (OSError, json.JSONDecodeError):
            return cls()
        if not isinstance(raw, dict):
            return cls()
        return cls.from_dict(raw)

    def save(self, path: Optional[Path] = None) -> Path:
        """Write settings JSON to disk, creating the parent dir if needed.

        Returns the path actually written.  Best-effort: if the file
        system refuses the write, the exception propagates — the GUI
        catches it and logs a warning so a read-only home dir doesn't
        crash the app on close.
        """
        p = path or default_settings_path()
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(
            json.dumps(self.to_dict(), indent=2, sort_keys=True),
            encoding="utf-8",
        )
        return p

    # --- Mutations with bounds ------------------------------------------- #

    _RECENT_CAP = 20

    def remember_inputs(self, paths: Iterable[Path]) -> None:
        """Move these paths to the top of ``recent_inputs`` (MRU order).

        Deduped via :func:`dedupe_paths` so filesystem-different spellings
        of the same file collapse into one entry.
        """
        resolved_new = [str(_safe_resolve(Path(p))) for p in paths]
        existing = [s for s in self.recent_inputs if s not in resolved_new]
        merged = resolved_new + existing
        self.recent_inputs = merged[: self._RECENT_CAP]


# ---------------------------------------------------------------------------
# Path helpers
# ---------------------------------------------------------------------------


def _safe_resolve(path: Path) -> Path:
    """``Path.resolve()`` that falls back to an absolute path if the
    file doesn't exist yet.  Needed on Windows where ``resolve(strict=False)``
    can still fail on some exotic paths; here we want 'best available'.
    """
    try:
        return path.resolve()
    except (OSError, RuntimeError):
        return path.absolute()


def dedupe_paths(paths: Iterable[Path]) -> List[Path]:
    """Return ``paths`` with duplicates removed, order preserved.

    Two paths are 'the same' if their resolved, absolute forms match —
    so ``./spec.docx``, ``spec.docx``, and ``../project/spec.docx`` (all
    pointing at the same file on disk) collapse to one entry.  This
    fixes REVIEW §2.11 where identity-based dedup let the same file slip
    into the input list twice via different spellings.
    """
    seen: set[Path] = set()
    out: List[Path] = []
    for p in paths:
        key = _safe_resolve(Path(p))
        if key in seen:
            continue
        seen.add(key)
        # Preserve the caller's original Path object rather than the
        # resolved form — the listbox looks nicer with the path the
        # user actually dragged/selected.
        out.append(Path(p))
    return out


def is_duplicate_of_any(candidate: Path, existing: Iterable[Path]) -> bool:
    """True if ``candidate`` resolves to the same file as any in ``existing``."""
    key = _safe_resolve(Path(candidate))
    return any(_safe_resolve(Path(p)) == key for p in existing)


# ---------------------------------------------------------------------------
# Actors-template generator
# ---------------------------------------------------------------------------


ACTORS_TEMPLATE_ROWS = [
    # (canonical, aliases)
    ("Auth Service", "Authentication Service, Auth"),
    ("Flight Software", "FSW, Onboard Software"),
    ("Ground Control", "GC, Ground Station, Ground Segment"),
    ("Payload Operator", "PL Op, Payload Op"),
]


def write_actors_template(output_path: Path) -> Path:
    """Write an example actors .xlsx to ``output_path``.

    The file matches the format ``load_actors_from_xlsx`` expects: a
    two-column sheet with a header row (Canonical name | Aliases).  It
    exists so a GUI user can hit "Save actors template…", open the file
    in Excel, replace the canned rows with their own, and hand it back
    to the extractor.  Fixes REVIEW §3.14.
    """
    # Lazy-import so the module stays usable without openpyxl (e.g. for
    # ``dedupe_paths`` consumers in tests).  openpyxl is already a core
    # dependency of the package so this import will almost always succeed.
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Actors"

    header_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Arial", size=10)

    # Header names must match what ``actors.load_actors_from_xlsx``
    # expects (case-insensitive): "Actor" and "Aliases".  Keep them
    # short so the generated template is immediately usable without any
    # header-row editing.
    headers = ("Actor", "Aliases")
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions[get_column_letter(1)].width = 24
    ws.column_dimensions[get_column_letter(2)].width = 52
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for row_idx, (canonical, aliases) in enumerate(ACTORS_TEMPLATE_ROWS, start=2):
        ws.cell(row=row_idx, column=1, value=canonical).font = body_font
        ws.cell(row=row_idx, column=2, value=aliases).font = body_font

    # A second sheet with a short usage note so the file is self-documenting.
    help_ws = wb.create_sheet("Readme")
    help_ws.column_dimensions["A"].width = 100
    notes = [
        "How to use this actors template:",
        "",
        "1. Replace the example rows in the 'Actors' sheet with your own.",
        "2. 'Canonical name' is the name the tool will report in the output.",
        "3. 'Aliases' is a comma-separated list of alternate spellings that",
        "   should all collapse to the canonical form (case-insensitive).",
        "4. Leave the header row alone (first row, column A + B).",
        "5. Save the file, then point the Requirements Extractor at it via",
        "   the 'Actors list' field.",
    ]
    for i, line in enumerate(notes, start=1):
        cell = help_ws.cell(row=i, column=1, value=line)
        cell.font = body_font
        if i == 1:
            cell.font = Font(name="Arial", size=11, bold=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
