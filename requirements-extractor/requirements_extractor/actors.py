"""Actor resolution.

Primary actor comes from the first column of the 2-column table row.
Secondary actors come from (in order of preference):
  1. A user-supplied actors list (Excel file with an "Actor" column, and an
     optional "Aliases" column containing comma-separated alternates).
  2. An optional spaCy-based NER pass (only if the user asked for it and
     spaCy is installed with an English model).
  3. An optional rule-based heuristics pass (the offline-network fallback;
     see :mod:`requirements_extractor.actor_heuristics`).
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator, List, Optional, Sequence, Tuple


@dataclass
class ActorEntry:
    name: str
    aliases: List[str]

    def all_forms(self) -> List[str]:
        forms = [self.name] + [a for a in self.aliases if a]
        return [f.strip() for f in forms if f and f.strip()]


# ---------------------------------------------------------------------------
# NER canonicalisation (REVIEW section 1.7)
#
# spaCy NER returns entity text verbatim, which in specs produces noisy
# surfaces like "the Auth Service" (leading determiner), "the Gateway's"
# (trailing possessive), or plain "ISO" / "USA" (ORG-labelled terms that
# aren't actors).  We tidy these up in one place so both the NLP pass in
# this file and any future external consumer can use the same rules.
# ---------------------------------------------------------------------------


#: Leading determiners to strip (lower-cased).  Order matters only for
#: documentation -- the matching is a straight startswith on each entry.
_LEADING_DETERMINERS: Tuple[str, ...] = ("the ", "a ", "an ")


#: Trailing possessive suffixes (ASCII apostrophe and curly apostrophe)
#: that should be removed so "Gateway's" becomes "Gateway".
_TRAILING_POSSESSIVES: Tuple[str, ...] = ("'s", "’s")


_WORD_TOKEN_RE = re.compile(r"\w+")


def canonicalise_ner_name(
    raw: str,
    *,
    canonical_names: Optional[Sequence[str]] = None,
) -> Optional[str]:
    """Clean up a raw NER entity string into an actor-shaped name.

    Strips leading determiners (``the``/``a``/``an``) and trailing
    possessives (``'s`` / curly-apostrophe ``’s``).  Returns
    ``None`` when the result would be empty or contains no letters/
    digits -- drop those rather than surface garbage to the reviewer.

    ``canonical_names`` is an optional whitelist.  When provided, a
    cleaned entity must share at least one token (case-insensitive,
    word-bounded) with at least one canonical name to survive.  This
    is the filter that rescues users from NLP noise like ``ISO`` /
    ``USA`` / ``Corp``: if those aren't in the user's curated actors
    list, they get dropped even though spaCy labelled them ORG.
    Without ``canonical_names`` (i.e. no user-supplied actors file),
    every cleaned entity passes through -- the caller is trusting the
    entity-label whitelist alone in that mode.
    """
    if not raw:
        return None
    s = raw.strip()
    # Strip leading determiners.  Done iteratively in case the entity
    # text happens to contain a second determiner after the first is
    # removed (rare but cheap to handle).
    while True:
        lowered = s.lower()
        stripped = False
        for det in _LEADING_DETERMINERS:
            if lowered.startswith(det):
                s = s[len(det):].strip()
                stripped = True
                break
        if not stripped:
            break
    # Strip trailing possessives.
    for poss in _TRAILING_POSSESSIVES:
        if s.endswith(poss):
            s = s[: -len(poss)].strip()
            break
    # Drop anything that's now empty or has no alphanumeric content.
    if not s or not any(c.isalnum() for c in s):
        return None
    # A bare determiner on its own (no following word) is garbage -- the
    # loop above only strips "the " (with trailing space), so "the"
    # alone slips through.  Catch it here.
    if s.lower() in {"the", "a", "an"}:
        return None
    # Canonical-overlap filter: token-level intersection with any name
    # in the canonical list.  Whole-word matching so "Auth" inside
    # "Authoring" doesn't falsely overlap.
    if canonical_names:
        s_tokens = {t.lower() for t in _WORD_TOKEN_RE.findall(s)}
        if not s_tokens:
            return None
        for canon in canonical_names:
            canon_tokens = {t.lower() for t in _WORD_TOKEN_RE.findall(canon or "")}
            if s_tokens & canon_tokens:
                return s
        return None
    return s


class ActorResolver:
    """Looks up secondary actors in a requirement's text.

    The resolver is deliberately tolerant -- it does case-insensitive
    word-boundary matching against every known name/alias and returns a
    deduped list of canonical names.
    """

    # Labels we treat as actor-ish when scanning NLP entities.
    #
    # REVIEW section 1.7: NORP (nationalities / religious / political
    # groups) and PRODUCT were both listed originally but produce noise
    # on specs -- "ISO"-as-NORP, "Windows"-as-PRODUCT, etc.  Narrowed
    # to the two labels that reliably mean "a named actor": people and
    # organisations.  The canonicalisation helper filters further by
    # overlap with the user-supplied actors list where one exists.
    _NLP_ACTOR_LABELS = frozenset({"PERSON", "ORG"})

    def __init__(
        self,
        actors: Optional[Sequence[ActorEntry]] = None,
        use_nlp: bool = False,
        use_heuristics: bool = False,
    ) -> None:
        self.actors: List[ActorEntry] = list(actors or [])
        self.use_nlp = use_nlp
        self.use_heuristics = use_heuristics
        self._nlp = None
        if use_nlp:
            self._nlp = _try_load_spacy()

        # Pre-build a single regex for all alias forms, mapped back to canonical
        # names.  This is fast even on large specs.
        self._alias_to_canonical: dict = {}
        patterns: List[str] = []
        for entry in self.actors:
            for form in entry.all_forms():
                key = form.lower()
                self._alias_to_canonical[key] = entry.name
                patterns.append(re.escape(form))
        if patterns:
            self._actor_re = re.compile(
                r"\b(" + "|".join(patterns) + r")\b",
                flags=re.IGNORECASE,
            )
        else:
            self._actor_re = None

    # --- Public API -----------------------------------------------------

    def has_nlp(self) -> bool:
        """Return True if an NLP pipeline is loaded and available."""
        return self._nlp is not None

    def iter_regex_hits(
        self, text: str, primary: str = "",
    ) -> Iterator[str]:
        """Yield canonical actor names found via alias-regex matching.

        Honors ``primary`` (excluded) and internally dedupes.  Callers
        wanting cross-source dedup should provide their own ``seen`` set
        via :meth:`iter_matches`.
        """
        if self._actor_re is None or not text:
            return
        primary_lower = (primary or "").strip().lower()
        seen: set = set()
        for m in self._actor_re.finditer(text):
            canonical = self._alias_to_canonical[m.group(1).lower()]
            key = canonical.lower()
            if key == primary_lower or key in seen:
                continue
            seen.add(key)
            yield canonical

    def iter_nlp_hits(
        self, text: str, primary: str = "",
    ) -> Iterator[str]:
        """Yield actor names found via NLP entity recognition.

        Returns nothing if NLP is unavailable or fails at runtime.
        Each candidate entity is passed through
        :func:`canonicalise_ner_name` to strip determiners and
        possessives, and (when a user-supplied actors list is
        present) to filter down to entities that share a token with
        at least one canonical name.
        """
        if self._nlp is None or not text:
            return
        primary_lower = (primary or "").strip().lower()
        seen: set = set()
        try:
            doc = self._nlp(text)
        except Exception:  # noqa: BLE001 -- NLP is best-effort
            return
        # Build the canonical-name list once per call.  When no user
        # actors list is loaded, ``canonical`` is None and the
        # canonicaliser skips the overlap filter entirely.
        canonical: Optional[List[str]] = None
        if self.actors:
            canonical = []
            for entry in self.actors:
                canonical.extend(entry.all_forms())
        for ent in doc.ents:
            if ent.label_ not in self._NLP_ACTOR_LABELS:
                continue
            cleaned = canonicalise_ner_name(ent.text, canonical_names=canonical)
            if cleaned is None:
                continue
            key = cleaned.lower()
            if key == primary_lower or key in seen:
                continue
            seen.add(key)
            yield cleaned

    def iter_heuristic_hits(
        self, text: str, primary: str = "",
    ) -> Iterator[str]:
        """Yield actor names from the rule-based heuristics fallback.

        Rule-based -- no NLP or seed list required. See
        :mod:`requirements_extractor.actor_heuristics` for the
        per-rule documentation. Returns nothing when
        ``use_heuristics=False`` (the default) so existing fixtures
        that depend on no-secondary-actor behaviour stay green.
        """
        if not self.use_heuristics or not text:
            return
        # Local import keeps the heuristics module out of the import
        # graph for callers that don't opt in -- the regex compile in
        # that module is non-trivial.
        from .actor_heuristics import extract_actor_candidates

        primary_lower = (primary or "").strip().lower()
        seen: set = set()
        for cand in extract_actor_candidates(text, primary=primary):
            key = cand.lower()
            if key == primary_lower or key in seen:
                continue
            seen.add(key)
            yield cand

    def iter_matches(
        self, text: str, primary: str = "",
    ) -> Iterator[Tuple[str, str]]:
        """Yield (name, source) tuples across all enabled passes.

        ``source`` is ``"regex"``, ``"nlp"``, or ``"rule"``.  Cross-
        source dedup is enforced: the same canonical name (case-
        insensitive) is emitted at most once regardless of which pass
        discovered it.  Order is regex -> nlp -> rule so the
        higher-confidence sources win cross-source attribution.
        """
        primary_lower = (primary or "").strip().lower()
        seen: set = set()
        for name in self.iter_regex_hits(text, primary):
            key = name.lower()
            if key == primary_lower or key in seen:
                continue
            seen.add(key)
            yield (name, "regex")
        for name in self.iter_nlp_hits(text, primary):
            key = name.lower()
            if key == primary_lower or key in seen:
                continue
            seen.add(key)
            yield (name, "nlp")
        for name in self.iter_heuristic_hits(text, primary):
            key = name.lower()
            if key == primary_lower or key in seen:
                continue
            seen.add(key)
            yield (name, "rule")

    def resolve(self, text: str, primary: str) -> List[str]:
        """Return a deduped list of secondary actor names found in ``text``."""
        return [name for name, _ in self.iter_matches(text, primary)]


def load_actors_from_xlsx(path: Path) -> List[ActorEntry]:
    """Load an actor list from an Excel file.

    Expected columns (header row, case-insensitive):
        Actor      -- canonical name (required)
        Aliases    -- comma-separated alternates (optional)
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c or "").strip().lower() for c in rows[0]]
    try:
        name_col = header.index("actor")
    except ValueError:
        raise ValueError(
            f"{path.name}: expected a header column named 'Actor' in row 1."
        )
    alias_col = header.index("aliases") if "aliases" in header else None

    entries: List[ActorEntry] = []
    for row in rows[1:]:
        if not row or row[name_col] is None:
            continue
        name = str(row[name_col]).strip()
        if not name:
            continue
        aliases: List[str] = []
        if alias_col is not None and row[alias_col] is not None:
            raw = str(row[alias_col])
            aliases = [a.strip() for a in raw.split(",") if a.strip()]
        entries.append(ActorEntry(name=name, aliases=aliases))
    return entries


def _try_load_spacy():
    """Return a loaded spaCy pipeline, or None if unavailable.

    ``spacy.load`` can raise ``OSError`` (model directory missing),
    ``ImportError`` (model package not installed), ``ValueError`` (old
    model incompatible with installed spacy), or -- in edge cases with
    spaCy's pydantic-based config -- ``TypeError`` when the pydantic
    major version mismatches.  All four are "spaCy isn't usable, move
    on".
    """
    try:
        import spacy  # type: ignore
    except ImportError:
        return None
    for model in ("en_core_web_sm", "en_core_web_md", "en_core_web_lg"):
        try:
            return spacy.load(model)
        except (ImportError, OSError, ValueError, TypeError):
            continue
    return None
