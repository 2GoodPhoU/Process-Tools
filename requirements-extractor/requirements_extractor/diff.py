"""Diff two extractor-produced workbooks.

REVIEW §3.12.  Given ``old.xlsx`` and ``new.xlsx`` (both produced by
:func:`requirements_extractor.writer.write_requirements`), emit a
third workbook highlighting:

* **Added** rows — present in ``new``, not in ``old`` (and not a
  paired change, see below).
* **Removed** rows — present in ``old``, not in ``new``.
* **Changed** rows — present at the same ``(source_file, row_ref)`` in
  both runs but with a different requirement ``text``.  This is the
  case that stable-ID matching misses: stable_id hashes
  ``(source_file, primary_actor, text)`` so any text edit produces a
  new ID.

The output has one header row plus one body row per diff entry, with
a coloured ``Change Type`` column so reviewers can skim.  A second
sheet summarises counts.  Works entirely on the xlsx files — no need
to re-parse the source .docx.

Pure module; no CLI wiring here.  ``requirements_extractor.cli``
exposes this via the ``diff`` subcommand.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Data model for a diff entry.
# ---------------------------------------------------------------------------


@dataclass
class DiffEntry:
    """One row in the diff output."""

    change_type: str            # "Added" | "Removed" | "Changed"
    stable_id: str
    source_file: str
    row_ref: str
    primary_actor: str
    text: str                   # New text (for Added / Changed) or old (for Removed)
    old_text: str = ""          # Populated for Changed rows only


#: Header fill / font for the output workbook — kept inline so this
#: module doesn't import from writer.py (which would be a circular
#: risk if writer.py ever wants to use the diff format).
_HEADER_FILL = PatternFill("solid", start_color="1F3864")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(name="Arial", size=10)

# Change-type fills: green for add, red for remove, yellow for change.
_ADDED_FILL = PatternFill("solid", start_color="D9EAD3")     # light green
_REMOVED_FILL = PatternFill("solid", start_color="F4CCCC")   # light red
_CHANGED_FILL = PatternFill("solid", start_color="FFF2CC")   # light yellow


# ---------------------------------------------------------------------------
# Reading a requirements workbook back into a dict.
# ---------------------------------------------------------------------------


def _read_requirements_workbook(path: Path) -> List[Dict[str, str]]:
    """Load the Requirements sheet as a list of {header: value} dicts.

    Accepts any workbook produced by
    :func:`writer.write_requirements`.  Prefers the sheet literally
    named ``Requirements`` (the writer's default title) over whatever
    was ``active`` when the file was last saved — otherwise a user who
    re-opened the file and left the Summary tab selected would cause
    the diff to silently read the wrong sheet.

    Raises ``ValueError`` if the workbook has no ``ID`` column in its
    header row.  Pre-stable-id outputs (tool versions before the
    stable-ID feature shipped) fall into that bucket — better to
    surface a clear error than silently emit a diff where every row
    is either Added or Removed.  Missing cells inside a row are
    tolerated and become empty strings.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    # Prefer the named sheet so a user-saved workbook with a different
    # active tab doesn't confuse the reader.
    ws = wb["Requirements"] if "Requirements" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c or "").strip() for c in rows[0]]
    if "ID" not in headers:
        raise ValueError(
            f"{Path(path).name}: no 'ID' column found in the header row. "
            "This workbook looks like it was produced by a tool version "
            "older than the stable-ID feature.  Re-run the extractor on "
            "the source .docx files to produce a diff-compatible xlsx."
        )
    out: List[Dict[str, str]] = []
    for row in rows[1:]:
        record: Dict[str, str] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = row[idx] if idx < len(row) else None
            record[header] = "" if value is None else str(value)
        # A row with no stable ID (shouldn't happen for modern outputs —
        # we already raised above if the column itself is missing) is
        # unusable for diff, skip it rather than produce a bogus entry.
        if not record.get("ID"):
            continue
        out.append(record)
    return out


# ---------------------------------------------------------------------------
# Diff computation.
# ---------------------------------------------------------------------------


def compute_diff(
    old_rows: Sequence[Dict[str, str]],
    new_rows: Sequence[Dict[str, str]],
) -> List[DiffEntry]:
    """Compute the diff between two requirements-workbook row lists.

    Matching strategy:

    1. Primary match on ``ID`` (stable_id) — exact stable_id collision
       means same file, actor, text → unchanged, not emitted.
    2. Secondary match on ``(Source File, Row Ref)`` for rows that
       didn't match on ID — catches the case where the author edited
       the requirement text at a stable position.  Emitted as
       ``Changed`` with both old and new text.
    3. Remaining old-only rows → ``Removed``.
    4. Remaining new-only rows → ``Added``.

    The ``(Source File, Row Ref)`` secondary match is intentionally
    narrow: if the row was reordered in the document, it'll look like
    Removed + Added rather than Changed.  That's a better failure
    mode than a false Changed pairing that would conflate two
    unrelated rows.
    """
    old_by_id = {r["ID"]: r for r in old_rows if r.get("ID")}
    new_by_id = {r["ID"]: r for r in new_rows if r.get("ID")}

    common_ids = old_by_id.keys() & new_by_id.keys()
    old_unmatched_ids = old_by_id.keys() - common_ids
    new_unmatched_ids = new_by_id.keys() - common_ids

    # Build a (file, row_ref) -> row lookup for the un-matched old set,
    # so we can detect "same position, new text" pairings.
    old_by_position: Dict[Tuple[str, str], Dict[str, str]] = {}
    for oid in old_unmatched_ids:
        r = old_by_id[oid]
        key = (r.get("Source File", ""), r.get("Row Ref", ""))
        # If two rows share a (file, row_ref) — e.g. multiple
        # requirements in one table row — only the first one slots in;
        # the others fall through to Removed.  Good enough for a first
        # pass; a future version could support multi-per-position.
        old_by_position.setdefault(key, r)

    entries: List[DiffEntry] = []

    # Pass 1: secondary-match detection — for each un-matched new row,
    # is there a same-position old row?
    paired_old_ids: set = set()
    for nid in sorted(new_unmatched_ids):
        r_new = new_by_id[nid]
        key = (r_new.get("Source File", ""), r_new.get("Row Ref", ""))
        r_old = old_by_position.get(key)
        if r_old is not None and r_old["ID"] not in paired_old_ids:
            entries.append(
                DiffEntry(
                    change_type="Changed",
                    stable_id=r_new.get("ID", ""),
                    source_file=r_new.get("Source File", ""),
                    row_ref=r_new.get("Row Ref", ""),
                    primary_actor=r_new.get("Primary Actor", ""),
                    text=r_new.get("Requirement", ""),
                    old_text=r_old.get("Requirement", ""),
                )
            )
            paired_old_ids.add(r_old["ID"])

    # Pass 2: Added — any new-unmatched row whose old counterpart
    # wasn't found by the secondary match.
    paired_new_positions = {
        (e.source_file, e.row_ref) for e in entries if e.change_type == "Changed"
    }
    for nid in sorted(new_unmatched_ids):
        r_new = new_by_id[nid]
        key = (r_new.get("Source File", ""), r_new.get("Row Ref", ""))
        if key in paired_new_positions:
            continue
        entries.append(
            DiffEntry(
                change_type="Added",
                stable_id=r_new.get("ID", ""),
                source_file=r_new.get("Source File", ""),
                row_ref=r_new.get("Row Ref", ""),
                primary_actor=r_new.get("Primary Actor", ""),
                text=r_new.get("Requirement", ""),
            )
        )

    # Pass 3: Removed — any old-unmatched row not paired into a Changed.
    for oid in sorted(old_unmatched_ids):
        if oid in paired_old_ids:
            continue
        r_old = old_by_id[oid]
        entries.append(
            DiffEntry(
                change_type="Removed",
                stable_id=r_old.get("ID", ""),
                source_file=r_old.get("Source File", ""),
                row_ref=r_old.get("Row Ref", ""),
                primary_actor=r_old.get("Primary Actor", ""),
                text=r_old.get("Requirement", ""),
            )
        )

    # Ordering: Removed → Changed → Added groups the output so reviewers
    # can scan each category, and within each group we sort by source
    # file then row ref for readability.
    order = {"Removed": 0, "Changed": 1, "Added": 2}
    entries.sort(key=lambda e: (
        order.get(e.change_type, 99), e.source_file, e.row_ref,
    ))
    return entries


# ---------------------------------------------------------------------------
# Writing the diff workbook.
# ---------------------------------------------------------------------------


_DIFF_COLUMNS = [
    ("Change Type", 12),
    ("ID", 14),
    ("Source File", 28),
    ("Row Ref", 18),
    ("Primary Actor", 22),
    ("Requirement", 70),
    ("Previous Text (for Changed)", 60),
]


def write_diff_workbook(
    entries: Sequence[DiffEntry],
    output_path: Path,
) -> Path:
    """Write the diff workbook to ``output_path``.

    Always creates a file — an empty diff produces a header-only sheet
    plus a summary noting "No changes" so scripts can tell the run
    completed vs. crashed.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Diff"

    for col_idx, (name, width) in enumerate(_DIFF_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    fill_by_type = {
        "Added": _ADDED_FILL,
        "Removed": _REMOVED_FILL,
        "Changed": _CHANGED_FILL,
    }

    for row_idx, entry in enumerate(entries, start=2):
        values = [
            entry.change_type,
            entry.stable_id,
            entry.source_file,
            entry.row_ref,
            entry.primary_actor,
            entry.text,
            entry.old_text if entry.change_type == "Changed" else "",
        ]
        fill = fill_by_type.get(entry.change_type)
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _BODY_FONT
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if fill is not None:
                cell.fill = fill

    # Summary sheet: count per change type.
    summary_ws = wb.create_sheet("Summary")
    summary_ws.column_dimensions["A"].width = 18
    summary_ws.column_dimensions["B"].width = 10
    summary_ws["A1"] = "Change Type"
    summary_ws["B1"] = "Count"
    summary_ws["A1"].font = _HEADER_FONT
    summary_ws["A1"].fill = _HEADER_FILL
    summary_ws["B1"].font = _HEADER_FONT
    summary_ws["B1"].fill = _HEADER_FILL
    counts = {"Added": 0, "Removed": 0, "Changed": 0}
    for entry in entries:
        counts[entry.change_type] = counts.get(entry.change_type, 0) + 1
    for i, change_type in enumerate(("Added", "Removed", "Changed"), start=2):
        summary_ws.cell(row=i, column=1, value=change_type).font = _BODY_FONT
        summary_ws.cell(row=i, column=2, value=counts[change_type]).font = _BODY_FONT

    wb.save(str(output_path))
    return output_path


# ---------------------------------------------------------------------------
# High-level entry point.
# ---------------------------------------------------------------------------


def diff_workbooks(
    old_path: Path, new_path: Path, output_path: Path,
) -> Tuple[Path, Dict[str, int]]:
    """End-to-end diff: read two workbooks, write the diff workbook.

    Returns a tuple of ``(output_path, counts)`` where ``counts`` is a
    dict with keys ``Added``, ``Removed``, ``Changed`` — convenient for
    CLI callers to print a one-line summary without re-reading the
    output.
    """
    old_rows = _read_requirements_workbook(Path(old_path))
    new_rows = _read_requirements_workbook(Path(new_path))
    entries = compute_diff(old_rows, new_rows)
    write_diff_workbook(entries, Path(output_path))
    counts = {"Added": 0, "Removed": 0, "Changed": 0}
    for e in entries:
        counts[e.change_type] = counts.get(e.change_type, 0) + 1
    return Path(output_path), counts
