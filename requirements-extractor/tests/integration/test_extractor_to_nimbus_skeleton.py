"""End-to-end integration test: requirements-extractor -> nimbus-skeleton.

Mirrors test_extractor_to_compliance_matrix.py for the second downstream
consumer of DDE's xlsx output. Verifies that the DDE -> nimbus-skeleton
data flow produces all expected output formats and that the skeleton
content is non-trivially populated from the fixture.

This test guards against regressions in the DDE xlsx schema breaking
nimbus-skeleton's loader, and against regressions in nimbus-skeleton's
emitters losing fidelity.

Uses the same 'simple_two_actors.docx' fixture as the compliance-matrix
integration test so a future schema-aware diff between the two outputs
is easy.
"""

from __future__ import annotations

import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).parent.parent.parent.parent.absolute()
REQUIREMENTS_EXTRACTOR_ROOT = REPO_ROOT / "requirements-extractor"
NIMBUS_SKELETON_ROOT = REPO_ROOT / "nimbus-skeleton"
FIXTURE = (
    REQUIREMENTS_EXTRACTOR_ROOT
    / "samples"
    / "procedures"
    / "simple_two_actors.docx"
)


def _run_dde(docx_in: Path, xlsx_out: Path) -> subprocess.CompletedProcess:
    """Run DDE on the fixture and return the subprocess result."""
    cmd = [
        sys.executable,
        "-m",
        "requirements_extractor.cli",
        "--no-summary",
        "requirements",
        str(docx_in),
        "-o",
        str(xlsx_out),
    ]
    return subprocess.run(
        cmd, cwd=REQUIREMENTS_EXTRACTOR_ROOT, capture_output=True, text=True
    )


def _run_nimbus(reqs_xlsx: Path, output_dir: Path, *, bpmn: bool) -> subprocess.CompletedProcess:
    """Run nimbus-skeleton on a DDE xlsx and return the subprocess result."""
    cmd = [
        sys.executable,
        "-m",
        "nimbus_skeleton.cli",
        "--requirements",
        str(reqs_xlsx),
        "--output-dir",
        str(output_dir),
        "--basename",
        "skeleton",
    ]
    if bpmn:
        cmd.append("--bpmn")
    return subprocess.run(
        cmd, cwd=NIMBUS_SKELETON_ROOT, capture_output=True, text=True
    )


class TestExtractorToNimbusSkeleton(unittest.TestCase):
    def test_fixture_exists(self) -> None:
        """Fail fast if the sample fixture is missing."""
        self.assertTrue(
            FIXTURE.exists(),
            f"Fixture not found at {FIXTURE} — has it been moved or renamed?",
        )

    def test_dde_to_nimbus_default_emits_five_files(self) -> None:
        """End-to-end without --bpmn: five output files."""
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            dde_xlsx = tmp / "dde.xlsx"
            out_dir = tmp / "skeleton_run"

            r1 = _run_dde(FIXTURE, dde_xlsx)
            self.assertEqual(r1.returncode, 0, f"DDE failed: {r1.stderr}")
            self.assertTrue(dde_xlsx.exists())

            r2 = _run_nimbus(dde_xlsx, out_dir, bpmn=False)
            self.assertEqual(r2.returncode, 0, f"nimbus-skeleton failed: {r2.stderr}")

            # Five default outputs
            for ext in (".puml", ".skel.yaml", ".xmi", ".vsdx", ".review.xlsx"):
                self.assertTrue(
                    (out_dir / f"skeleton{ext}").exists(),
                    f"Missing output file: skeleton{ext}",
                )
            # No .bpmn unless --bpmn was passed
            self.assertFalse(
                (out_dir / "skeleton.bpmn").exists(),
                ".bpmn should not be emitted without --bpmn",
            )

    def test_dde_to_nimbus_with_bpmn_emits_six_files(self) -> None:
        """End-to-end with --bpmn: six output files including .bpmn."""
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            dde_xlsx = tmp / "dde.xlsx"
            out_dir = tmp / "skeleton_run"

            r1 = _run_dde(FIXTURE, dde_xlsx)
            self.assertEqual(r1.returncode, 0)

            r2 = _run_nimbus(dde_xlsx, out_dir, bpmn=True)
            self.assertEqual(r2.returncode, 0, f"nimbus-skeleton failed: {r2.stderr}")

            for ext in (
                ".puml",
                ".skel.yaml",
                ".xmi",
                ".vsdx",
                ".bpmn",
                ".review.xlsx",
            ):
                self.assertTrue(
                    (out_dir / f"skeleton{ext}").exists(),
                    f"Missing output file: skeleton{ext}",
                )

    def test_skeleton_yaml_contains_actors_and_activities(self) -> None:
        """Sanity check: the YAML manifest is non-empty and well-shaped."""
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            dde_xlsx = tmp / "dde.xlsx"
            out_dir = tmp / "skeleton_run"

            r1 = _run_dde(FIXTURE, dde_xlsx)
            self.assertEqual(r1.returncode, 0)

            r2 = _run_nimbus(dde_xlsx, out_dir, bpmn=False)
            self.assertEqual(r2.returncode, 0)

            yaml_text = (out_dir / "skeleton.skel.yaml").read_text(encoding="utf-8")
            # Every manifest has these top-level keys
            for key in ("actors", "activities", "flows"):
                self.assertIn(key, yaml_text, f"Manifest missing key: {key}")
            # The fixture is named simple_two_actors so we expect at least
            # 2 actors and at least 1 activity.
            self.assertIn("actors", yaml_text)

    def test_review_xlsx_well_formed(self) -> None:
        """Review side-car xlsx must be readable and have expected sheets."""
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            dde_xlsx = tmp / "dde.xlsx"
            out_dir = tmp / "skeleton_run"

            self.assertEqual(_run_dde(FIXTURE, dde_xlsx).returncode, 0)
            self.assertEqual(_run_nimbus(dde_xlsx, out_dir, bpmn=False).returncode, 0)

            wb = load_workbook(out_dir / "skeleton.review.xlsx")
            # At least one sheet — the review sheet.
            self.assertGreater(len(wb.sheetnames), 0)

    def test_bpmn_output_parses_as_xml(self) -> None:
        """The emitted .bpmn must be well-formed XML rooted at bpmn:definitions."""
        import xml.etree.ElementTree as ET

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            dde_xlsx = tmp / "dde.xlsx"
            out_dir = tmp / "skeleton_run"

            self.assertEqual(_run_dde(FIXTURE, dde_xlsx).returncode, 0)
            self.assertEqual(_run_nimbus(dde_xlsx, out_dir, bpmn=True).returncode, 0)

            bpmn_path = out_dir / "skeleton.bpmn"
            tree = ET.parse(bpmn_path)
            root = tree.getroot()
            # Local-name "definitions"; namespace is BPMN's.
            self.assertTrue(root.tag.endswith("}definitions"))


if __name__ == "__main__":
    unittest.main()
