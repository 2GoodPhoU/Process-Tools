"""End-to-end integration test: requirements-extractor -> compliance-matrix.

Tests the data flow from Document Data Extractor (DDE) producing an xlsx
workbook through compliance-matrix consuming that workbook and producing a
coverage matrix.

This test verifies:
1. DDE can extract requirements from a fixture .docx
2. The output matches the expected xlsx schema (process-tools-common)
3. compliance-matrix can accept the DDE output
4. A coverage matrix is produced with expected structure

Uses 'simple_two_actors.docx' fixture (substantive procedure with requirements).
"""

import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).parent.parent.parent.parent.absolute()
REQUIREMENTS_EXTRACTOR_ROOT = REPO_ROOT / "requirements-extractor"
COMPLIANCE_MATRIX_ROOT = REPO_ROOT / "compliance-matrix"
FIXTURES_DIR = REQUIREMENTS_EXTRACTOR_ROOT / "samples" / "procedures"


def test_extractor_output_schema() -> None:
    """Phase 1: DDE produces xlsx with correct structure.

    Runs requirements-extractor on the fixture and verifies:
    - Output file exists
    - File is valid xlsx (readable by openpyxl)
    - Contains 'Requirements' sheet (per DDE spec)
    - 'Requirements' sheet has expected columns (ID, Requirement, etc.)
    - Row count > 0 (fixture produces actual requirements)
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        temp_dir_path = Path(tmpdir)
        fixture_docx = FIXTURES_DIR / "simple_two_actors.docx"
        output_xlsx = temp_dir_path / "dde_output.xlsx"

        # Run DDE via CLI
        cmd = [
            sys.executable,
            "-m",
            "requirements_extractor.cli",
            "--no-summary",
            "requirements",
            str(fixture_docx),
            "-o",
            str(output_xlsx),
        ]

        result = subprocess.run(cmd, cwd=REQUIREMENTS_EXTRACTOR_ROOT, capture_output=True, text=True)

        assert result.returncode == 0, f"DDE failed: {result.stderr}"
        assert output_xlsx.exists(), "DDE did not produce output xlsx"

        # Verify xlsx structure
        wb = load_workbook(output_xlsx)
        assert "Requirements" in wb.sheetnames, f"Missing 'Requirements' sheet. Found: {wb.sheetnames}"

        req_sheet = wb["Requirements"]
        assert req_sheet.max_row > 1, "Requirements sheet has no data rows"

        # Check expected columns (header row)
        headers = [cell.value for cell in req_sheet[1]]
        expected_cols = {"ID", "Requirement"}  # Core columns
        found_cols = {h for h in headers if h}
        assert expected_cols.issubset(found_cols), (
            f"Missing columns. Expected {expected_cols}, found {found_cols}"
        )

        # Verify data rows exist and have content
        data_rows = req_sheet.max_row - 1  # Exclude header
        assert data_rows > 0, f"No data rows extracted (max_row={req_sheet.max_row})"


def test_compliance_matrix_intake() -> None:
    """Phase 2: compliance-matrix accepts DDE output and produces matrix.

    Runs requirements-extractor on the fixture, then feeds the output
    to compliance-matrix (with the same xlsx as both contract and procedure
    for this smoke test) and verifies:
    - compliance-matrix exits cleanly
    - Output matrix xlsx exists
    - Contains 'Matrix' sheet (per compliance-matrix spec)
    - Matrix sheet has data (rows and columns)
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        temp_dir_path = Path(tmpdir)
        fixture_docx = FIXTURES_DIR / "simple_two_actors.docx"
        dde_output = temp_dir_path / "dde_output.xlsx"
        matrix_output = temp_dir_path / "matrix.xlsx"

        # Step 1: Run DDE
        cmd_dde = [
            sys.executable,
            "-m",
            "requirements_extractor.cli",
            "--no-summary",
            "requirements",
            str(fixture_docx),
            "-o",
            str(dde_output),
        ]
        result_dde = subprocess.run(cmd_dde, cwd=REQUIREMENTS_EXTRACTOR_ROOT, capture_output=True, text=True)
        assert result_dde.returncode == 0, f"DDE failed: {result_dde.stderr}"

        # Step 2: Run compliance-matrix
        # Use the same xlsx for both sides (self-comparison) as a smoke test
        cmd_cm = [
            sys.executable,
            "-m",
            "compliance_matrix.cli",
            "--contract",
            str(dde_output),
            "--procedure",
            str(dde_output),
            "-o",
            str(matrix_output),
        ]
        result_cm = subprocess.run(cmd_cm, cwd=COMPLIANCE_MATRIX_ROOT, capture_output=True, text=True)
        assert result_cm.returncode == 0, f"compliance-matrix failed: {result_cm.stderr}"
        assert matrix_output.exists(), "compliance-matrix did not produce output xlsx"

        # Verify matrix structure
        wb = load_workbook(matrix_output)
        assert "Matrix" in wb.sheetnames, f"Missing 'Matrix' sheet. Found: {wb.sheetnames}"

        matrix_sheet = wb["Matrix"]
        assert matrix_sheet.max_row > 1, "Matrix sheet has no data rows"
        assert matrix_sheet.max_column > 1, "Matrix sheet has no data columns"


def test_integration_data_flow() -> None:
    """Phase 3: End-to-end data flow with realistic setup.

    Extracts requirements from fixture, then treats them as both
    contract and procedure for matrix generation (self-consistency check).
    This is a simplified smoke test; real usage would have distinct
    contract and procedure documents.

    Verifies:
    - All expected sheets present in matrix output
    - Detail sheet exists and has data
    - Scores are numeric and in expected range [0, 1]
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        temp_dir_path = Path(tmpdir)
        fixture_docx = FIXTURES_DIR / "simple_two_actors.docx"
        dde_output = temp_dir_path / "dde_output.xlsx"
        matrix_output = temp_dir_path / "matrix.xlsx"

        # Run DDE
        cmd_dde = [
            sys.executable,
            "-m",
            "requirements_extractor.cli",
            "--no-summary",
            "requirements",
            str(fixture_docx),
            "-o",
            str(dde_output),
        ]
        result_dde = subprocess.run(cmd_dde, cwd=REQUIREMENTS_EXTRACTOR_ROOT, capture_output=True, text=True)
        assert result_dde.returncode == 0

        # Run compliance-matrix
        cmd_cm = [
            sys.executable,
            "-m",
            "compliance_matrix.cli",
            "--contract",
            str(dde_output),
            "--procedure",
            str(dde_output),
            "-o",
            str(matrix_output),
        ]
        result_cm = subprocess.run(cmd_cm, cwd=COMPLIANCE_MATRIX_ROOT, capture_output=True, text=True)
        assert result_cm.returncode == 0

        # Check all expected sheets
        wb = load_workbook(matrix_output)
        expected_sheets = {"Matrix", "Detail", "Gaps"}
        for sheet in expected_sheets:
            assert sheet in wb.sheetnames, f"Missing sheet: {sheet}"

        # Detail sheet should have data (matcher results)
        detail_sheet = wb["Detail"]
        assert detail_sheet.max_row > 1, "Detail sheet has no matched pairs"

        # Spot-check score format (should be numeric, in [0, 1])
        for row in list(detail_sheet.iter_rows(min_row=2, max_row=min(10, detail_sheet.max_row))):
            # Find numeric columns (likely scores)
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    assert 0 <= cell.value <= 1, f"Score out of range: {cell.value}"


if __name__ == "__main__":
    print("Running integration tests...")
    test_extractor_output_schema()
    print("✓ test_extractor_output_schema PASSED")
    test_compliance_matrix_intake()
    print("✓ test_compliance_matrix_intake PASSED")
    test_integration_data_flow()
    print("✓ test_integration_data_flow PASSED")
    print("\n✓✓✓ ALL INTEGRATION TESTS PASSED ✓✓✓")
