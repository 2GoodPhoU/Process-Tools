"""Tests for the stable-ID scheme and end-to-end dry-run behaviour.

These cover three layers:

1. :func:`compute_stable_id` purity — determinism, whitespace/case
   normalisation, and input-sensitivity to the three identity fields.
2. :func:`ensure_unique_stable_ids` — collision-suffixing behaviour on
   requirements with identical (file, actor, text).
3. The ID column showing up in the xlsx writer, and ``--dry-run`` in the
   CLI correctly skipping file writes while still exercising the rest
   of the pipeline.

Run:  python -m unittest tests.test_stable_ids
"""

from __future__ import annotations

import io
import sys
import tempfile
import unittest
from contextlib import redirect_stdout
from dataclasses import replace
from pathlib import Path

from openpyxl import load_workbook

from requirements_extractor.cli import main
from requirements_extractor.extractor import extract_from_files
from requirements_extractor.models import (
    Requirement,
    annotate_cross_source_duplicates,
    compute_stable_id,
    ensure_unique_stable_ids,
)


ROOT = Path(__file__).resolve().parent.parent
SAMPLE = ROOT / "samples" / "sample_spec.docx"


def _make_req(
    text: str,
    *,
    source_file: str = "spec.docx",
    primary_actor: str = "User",
    order: int = 1,
) -> Requirement:
    return Requirement(
        order=order,
        source_file=source_file,
        heading_trail="",
        section_topic="",
        row_ref="",
        block_ref="",
        primary_actor=primary_actor,
        secondary_actors=[],
        text=text,
        req_type="Hard",
        keywords=[],
        confidence="High",
        stable_id=compute_stable_id(source_file, primary_actor, text),
    )


# ---------------------------------------------------------------------------
# compute_stable_id — purity
# ---------------------------------------------------------------------------


class TestComputeStableId(unittest.TestCase):
    def test_format_is_req_prefix_plus_8_hex(self) -> None:
        sid = compute_stable_id("spec.docx", "User", "The User shall log in.")
        self.assertTrue(sid.startswith("REQ-"))
        # "REQ-" + 8 hex characters
        self.assertEqual(len(sid), 4 + 8)
        hex_part = sid[4:]
        int(hex_part, 16)  # raises ValueError if not valid hex

    def test_deterministic_across_calls(self) -> None:
        a = compute_stable_id("spec.docx", "User", "Text A")
        b = compute_stable_id("spec.docx", "User", "Text A")
        self.assertEqual(a, b)

    def test_whitespace_normalised(self) -> None:
        a = compute_stable_id("spec.docx", "User", "The User shall log in.")
        b = compute_stable_id("spec.docx", "User", "  The  User   shall  log  in.  ")
        self.assertEqual(a, b)

    def test_case_normalised(self) -> None:
        # Same actor with different capitalisation should not churn ID.
        a = compute_stable_id("spec.docx", "User", "text")
        b = compute_stable_id("spec.docx", "USER", "TEXT")
        self.assertEqual(a, b)

    def test_distinct_text_distinct_id(self) -> None:
        a = compute_stable_id("spec.docx", "User", "Text A")
        b = compute_stable_id("spec.docx", "User", "Text B")
        self.assertNotEqual(a, b)

    def test_distinct_file_distinct_id(self) -> None:
        a = compute_stable_id("one.docx", "User", "Same text")
        b = compute_stable_id("two.docx", "User", "Same text")
        self.assertNotEqual(a, b)

    def test_distinct_actor_distinct_id(self) -> None:
        a = compute_stable_id("spec.docx", "User",  "Same text")
        b = compute_stable_id("spec.docx", "Admin", "Same text")
        self.assertNotEqual(a, b)


# ---------------------------------------------------------------------------
# ensure_unique_stable_ids — collision suffixing
# ---------------------------------------------------------------------------


class TestEnsureUniqueStableIds(unittest.TestCase):
    def test_no_duplicates_left_alone(self) -> None:
        reqs = [
            _make_req("alpha"),
            _make_req("bravo"),
            _make_req("charlie"),
        ]
        originals = [r.stable_id for r in reqs]
        ensure_unique_stable_ids(reqs)
        self.assertEqual([r.stable_id for r in reqs], originals)

    def test_duplicates_get_suffixes_in_order(self) -> None:
        reqs = [
            _make_req("same", order=1),
            _make_req("same", order=2),
            _make_req("same", order=3),
            _make_req("other", order=4),
        ]
        base = reqs[0].stable_id
        ensure_unique_stable_ids(reqs)
        self.assertEqual(reqs[0].stable_id, base)
        self.assertEqual(reqs[1].stable_id, f"{base}-1")
        self.assertEqual(reqs[2].stable_id, f"{base}-2")
        # Unrelated row is untouched.
        self.assertEqual(reqs[3].stable_id, _make_req("other").stable_id)

    def test_empty_input_is_safe(self) -> None:
        ensure_unique_stable_ids([])  # no exception

    def test_empty_stable_id_skipped(self) -> None:
        # A requirement that somehow arrived without a stable_id (e.g.
        # constructed in an old test) should not be rewritten to "-1".
        reqs = [
            _make_req("alpha"),
            replace(_make_req("alpha"), stable_id=""),
        ]
        ensure_unique_stable_ids(reqs)
        self.assertNotEqual(reqs[0].stable_id, "")
        self.assertEqual(reqs[1].stable_id, "")


# ---------------------------------------------------------------------------
# Writer integration — ID column is present and populated
# ---------------------------------------------------------------------------


class TestWriterIdColumn(unittest.TestCase):
    def test_id_column_present_and_populated(self) -> None:
        if not SAMPLE.exists():  # pragma: no cover — sample is checked in
            self.skipTest("sample_spec.docx missing")

        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "out.xlsx"
            result = extract_from_files(
                input_paths=[SAMPLE], output_path=out,
            )
            self.assertTrue(out.exists())
            self.assertGreater(len(result.requirements), 0)

            wb = load_workbook(out)
            ws = wb.active
            headers = [c.value for c in ws[1]]
            self.assertIn("ID", headers)
            # ID should be the second column (immediately after #).
            self.assertEqual(headers[0], "#")
            self.assertEqual(headers[1], "ID")

            # Every body row's ID cell is a REQ- value.
            id_col = headers.index("ID") + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                    min_col=id_col, max_col=id_col,
                                    values_only=True):
                (value,) = row
                self.assertIsInstance(value, str)
                self.assertTrue(value.startswith("REQ-"), value)

    def test_ids_stable_across_runs_on_same_input(self) -> None:
        if not SAMPLE.exists():  # pragma: no cover
            self.skipTest("sample_spec.docx missing")

        with tempfile.TemporaryDirectory() as d:
            a = extract_from_files([SAMPLE], Path(d) / "a.xlsx")
            b = extract_from_files([SAMPLE], Path(d) / "b.xlsx")
        ids_a = [r.stable_id for r in a.requirements]
        ids_b = [r.stable_id for r in b.requirements]
        self.assertEqual(ids_a, ids_b)
        self.assertTrue(all(sid.startswith("REQ-") for sid in ids_a))


# ---------------------------------------------------------------------------
# --dry-run end-to-end
# ---------------------------------------------------------------------------


class TestDryRunEndToEnd(unittest.TestCase):
    def test_dry_run_writes_no_xlsx(self) -> None:
        if not SAMPLE.exists():  # pragma: no cover
            self.skipTest("sample_spec.docx missing")

        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "should_not_exist.xlsx"
            result = extract_from_files(
                input_paths=[SAMPLE], output_path=out, dry_run=True,
            )
            self.assertTrue(result.dry_run)
            self.assertIsNone(result.output_path)
            self.assertIsNone(result.statement_set_path)
            self.assertFalse(out.exists())
            # But parsing happened — we got requirements and IDs.
            self.assertGreater(len(result.requirements), 0)
            self.assertTrue(all(r.stable_id for r in result.requirements))

    def test_dry_run_skips_statement_set(self) -> None:
        if not SAMPLE.exists():  # pragma: no cover
            self.skipTest("sample_spec.docx missing")

        with tempfile.TemporaryDirectory() as d:
            stmt = Path(d) / "ss.csv"
            result = extract_from_files(
                input_paths=[SAMPLE],
                output_path=Path(d) / "req.xlsx",
                statement_set_path=stmt,
                dry_run=True,
            )
            self.assertTrue(result.dry_run)
            self.assertFalse(stmt.exists())
            self.assertIsNone(result.statement_set_path)


# ---------------------------------------------------------------------------
# CLI wiring for --dry-run / --show-samples
# ---------------------------------------------------------------------------


class TestDryRunCLI(unittest.TestCase):
    def test_dry_run_flag_parsed(self) -> None:
        from requirements_extractor.cli import build_parser
        args = build_parser().parse_args([
            "requirements", "spec.docx", "--dry-run",
        ])
        self.assertTrue(args.dry_run)
        self.assertEqual(args.show_samples, 0)

    def test_show_samples_flag_parsed(self) -> None:
        from requirements_extractor.cli import build_parser
        args = build_parser().parse_args([
            "requirements", "spec.docx", "--show-samples", "3",
        ])
        self.assertEqual(args.show_samples, 3)

    def test_dry_run_default_false(self) -> None:
        from requirements_extractor.cli import build_parser
        args = build_parser().parse_args(["requirements", "spec.docx"])
        self.assertFalse(args.dry_run)
        self.assertEqual(args.show_samples, 0)

    def test_main_dry_run_end_to_end(self) -> None:
        if not SAMPLE.exists():  # pragma: no cover
            self.skipTest("sample_spec.docx missing")

        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "out.xlsx"
            buf = io.StringIO()
            with redirect_stdout(buf):
                rc = main([
                    "requirements", str(SAMPLE),
                    "-o", str(out),
                    "--dry-run",
                    "--show-samples", "2",
                ])
            self.assertEqual(rc, 0)
            self.assertFalse(out.exists())
            output = buf.getvalue()
            self.assertIn("dry run", output)
            # Two sample lines expected.
            self.assertIn("REQ-", output)
            self.assertIn("First 2 sample(s):", output)


# ---------------------------------------------------------------------------
# annotate_cross_source_duplicates — REVIEW §1.10
# ---------------------------------------------------------------------------


class TestCrossSourceDedup(unittest.TestCase):
    """Flag ``(actor, text)`` duplicates across files by annotating the
    later row's notes with a pointer to the original.
    """

    def test_unique_rows_untouched(self) -> None:
        reqs = [
            _make_req("Text A", source_file="a.docx"),
            _make_req("Text B", source_file="b.docx"),
        ]
        flagged = annotate_cross_source_duplicates(reqs)
        self.assertEqual(flagged, 0)
        for r in reqs:
            self.assertEqual(r.notes, "")

    def test_cross_file_duplicate_annotated(self) -> None:
        """Same sentence in two different source files — the second
        row should pick up a 'Duplicate of <stable_id> (<file>, <row>)'
        note."""
        r_a = _make_req(
            "The User shall authenticate.",
            source_file="a.docx",
            primary_actor="User",
        )
        r_a.row_ref = "Table 1, Row 3"
        r_b = _make_req(
            "The User shall authenticate.",
            source_file="b.docx",
            primary_actor="User",
        )
        r_b.row_ref = "Table 2, Row 1"
        flagged = annotate_cross_source_duplicates([r_a, r_b])
        self.assertEqual(flagged, 1)
        self.assertEqual(r_a.notes, "")
        self.assertIn("Duplicate of", r_b.notes)
        self.assertIn(r_a.stable_id, r_b.notes)
        self.assertIn("a.docx", r_b.notes)
        self.assertIn("Table 1, Row 3", r_b.notes)

    def test_different_actors_not_duplicates(self) -> None:
        # Same text but different primary actors — NOT a duplicate.
        # Actor is part of the identity.
        r_a = _make_req(
            "The target shall be acquired.",
            primary_actor="Operator",
        )
        r_b = _make_req(
            "The target shall be acquired.",
            primary_actor="Supervisor",
        )
        flagged = annotate_cross_source_duplicates([r_a, r_b])
        self.assertEqual(flagged, 0)
        self.assertEqual(r_b.notes, "")

    def test_whitespace_and_case_insensitive(self) -> None:
        # Cosmetic differences must not prevent a match — otherwise
        # formatting-only edits would mask a real duplicate.
        r_a = _make_req("The User shall authenticate.", source_file="a.docx")
        r_b = _make_req(
            "  The  user   SHALL authenticate.  ",
            source_file="b.docx",
        )
        flagged = annotate_cross_source_duplicates([r_a, r_b])
        self.assertEqual(flagged, 1)
        self.assertIn("Duplicate of", r_b.notes)

    def test_preserves_existing_notes(self) -> None:
        r_a = _make_req("Text X", source_file="a.docx")
        r_b = _make_req("Text X", source_file="b.docx")
        r_b.notes = "Soft language — verify with author."
        annotate_cross_source_duplicates([r_a, r_b])
        self.assertIn("Soft language", r_b.notes)
        self.assertIn("Duplicate of", r_b.notes)
        self.assertGreaterEqual(r_b.notes.count("\n"), 1)

    def test_multi_way_duplicate_points_to_first(self) -> None:
        # Three rows with the same (actor, text).  Rows 2 and 3 both
        # point back to Row 1, not to each other.
        reqs = [
            _make_req("Text Y", source_file="a.docx"),
            _make_req("Text Y", source_file="b.docx"),
            _make_req("Text Y", source_file="c.docx"),
        ]
        flagged = annotate_cross_source_duplicates(reqs)
        self.assertEqual(flagged, 2)
        self.assertEqual(reqs[0].notes, "")
        for later in reqs[1:]:
            self.assertIn(reqs[0].stable_id, later.notes)

    def test_empty_text_rows_ignored(self) -> None:
        # Defensive: an empty-text row must not poison the seen-map
        # for every subsequent empty-text row.
        reqs = [
            _make_req("", source_file="a.docx"),
            _make_req("", source_file="b.docx"),
        ]
        flagged = annotate_cross_source_duplicates(reqs)
        self.assertEqual(flagged, 0)


if __name__ == "__main__":
    unittest.main()
