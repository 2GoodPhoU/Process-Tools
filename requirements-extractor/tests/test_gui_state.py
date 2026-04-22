"""Headless tests for requirements_extractor.gui_state.

These exercise the Tk-free helpers (settings persistence, path dedup,
actors-template generation) directly so they run in any CI environment
without needing a display.

Run:  python -m unittest tests.test_gui_state
"""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from requirements_extractor.gui_state import (
    ACTORS_TEMPLATE_ROWS,
    GuiSettings,
    dedupe_paths,
    default_settings_path,
    is_duplicate_of_any,
    write_actors_template,
)


class TestDefaultSettingsPath(unittest.TestCase):
    def test_lives_under_home(self) -> None:
        p = default_settings_path()
        self.assertEqual(p.name, "settings.json")
        # Parent folder is versionable by name; make sure the location
        # isn't accidentally pointing at the CWD.
        self.assertTrue(str(p).startswith(str(Path.home())))


class TestGuiSettingsDefaults(unittest.TestCase):
    def test_defaults_are_safe(self) -> None:
        s = GuiSettings()
        self.assertEqual(s.schema_version, 1)
        self.assertEqual(s.window_geometry, "760x560")
        self.assertEqual(s.last_actors_path, "")
        self.assertEqual(s.last_output_path, "")
        self.assertEqual(s.last_input_dir, "")
        self.assertFalse(s.use_nlp)
        self.assertFalse(s.export_statement_set)
        self.assertTrue(s.open_output_on_done)
        self.assertFalse(s.dry_run)
        self.assertEqual(s.mode, "requirements")
        self.assertEqual(s.recent_inputs, [])


class TestGuiSettingsMode(unittest.TestCase):
    def test_known_modes_round_trip(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            path = Path(d) / "s.json"
            for m in ("requirements", "actors"):
                GuiSettings(mode=m).save(path)
                self.assertEqual(GuiSettings.load(path).mode, m)

    def test_unknown_mode_on_disk_falls_back(self) -> None:
        """A hand-edited or future-version settings file with a bogus
        mode value must not wedge the GUI."""
        with tempfile.TemporaryDirectory() as d:
            path = Path(d) / "s.json"
            path.write_text(
                json.dumps({"mode": "quantum"}), encoding="utf-8",
            )
            self.assertEqual(GuiSettings.load(path).mode, "requirements")


class TestGuiSettingsRoundtrip(unittest.TestCase):
    def test_save_and_reload_preserves_values(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            target = Path(d) / "nested" / "settings.json"  # parent must auto-create
            before = GuiSettings(
                window_geometry="900x700+10+20",
                last_actors_path="/tmp/actors.xlsx",
                last_output_path="/tmp/out.xlsx",
                use_nlp=True,
                export_statement_set=True,
                open_output_on_done=False,
                dry_run=True,
                recent_inputs=["/tmp/a.docx", "/tmp/b.docx"],
            )
            written = before.save(target)
            self.assertEqual(written, target)
            self.assertTrue(target.exists())

            after = GuiSettings.load(target)
            self.assertEqual(after.window_geometry, "900x700+10+20")
            self.assertEqual(after.last_actors_path, "/tmp/actors.xlsx")
            self.assertEqual(after.last_output_path, "/tmp/out.xlsx")
            self.assertTrue(after.use_nlp)
            self.assertTrue(after.export_statement_set)
            self.assertFalse(after.open_output_on_done)
            self.assertTrue(after.dry_run)
            self.assertEqual(after.recent_inputs, ["/tmp/a.docx", "/tmp/b.docx"])

    def test_load_missing_file_returns_defaults(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            missing = Path(d) / "nope.json"
            loaded = GuiSettings.load(missing)
            # Should deep-equal a fresh defaults instance.
            self.assertEqual(loaded, GuiSettings())

    def test_load_malformed_json_returns_defaults(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            path = Path(d) / "bad.json"
            path.write_text("not json at all { {{", encoding="utf-8")
            self.assertEqual(GuiSettings.load(path), GuiSettings())

    def test_load_non_dict_json_returns_defaults(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            path = Path(d) / "list.json"
            path.write_text("[1, 2, 3]", encoding="utf-8")
            self.assertEqual(GuiSettings.load(path), GuiSettings())

    def test_unknown_keys_are_ignored(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            path = Path(d) / "future.json"
            path.write_text(
                json.dumps({
                    "window_geometry": "100x100",
                    "this_field_does_not_exist": True,
                    "another_unknown": [1, 2, 3],
                }),
                encoding="utf-8",
            )
            loaded = GuiSettings.load(path)
            self.assertEqual(loaded.window_geometry, "100x100")
            # Known defaults still apply.
            self.assertEqual(loaded.schema_version, 1)

    def test_type_mismatch_falls_back_to_default(self) -> None:
        """A boolean field receiving a string should not crash or corrupt."""
        with tempfile.TemporaryDirectory() as d:
            path = Path(d) / "wrong_types.json"
            path.write_text(
                json.dumps({
                    "use_nlp": "yes please",  # wrong type
                    "last_output_path": 42,    # wrong type
                    "recent_inputs": "not a list",  # wrong type
                    "last_actors_path": "/good/path",  # right type — kept
                }),
                encoding="utf-8",
            )
            loaded = GuiSettings.load(path)
            self.assertFalse(loaded.use_nlp)  # default
            self.assertEqual(loaded.last_output_path, "")  # default
            self.assertEqual(loaded.recent_inputs, [])  # default
            self.assertEqual(loaded.last_actors_path, "/good/path")  # preserved


class TestRememberInputs(unittest.TestCase):
    def test_mru_ordering(self) -> None:
        s = GuiSettings()
        s.remember_inputs([Path("/tmp/a.docx"), Path("/tmp/b.docx")])
        self.assertEqual(
            s.recent_inputs,
            [str(Path("/tmp/a.docx").resolve()), str(Path("/tmp/b.docx").resolve())],
        )
        # Adding b again should move it to the top, not duplicate.
        s.remember_inputs([Path("/tmp/b.docx")])
        self.assertEqual(
            s.recent_inputs,
            [str(Path("/tmp/b.docx").resolve()), str(Path("/tmp/a.docx").resolve())],
        )

    def test_cap_enforced(self) -> None:
        s = GuiSettings()
        # Force the cap down to 3 so the test stays fast and readable.
        s._RECENT_CAP = 3  # type: ignore[attr-defined]
        s.remember_inputs([Path(f"/tmp/{i}.docx") for i in range(10)])
        self.assertEqual(len(s.recent_inputs), 3)


class TestPathDedup(unittest.TestCase):
    def test_identical_paths_dedupe(self) -> None:
        out = dedupe_paths([Path("/tmp/a.docx"), Path("/tmp/a.docx")])
        self.assertEqual(len(out), 1)

    def test_different_spellings_of_same_file_dedupe(self) -> None:
        """REVIEW §2.11 regression — ./x/../x/a.docx == x/a.docx."""
        with tempfile.TemporaryDirectory() as d:
            real = Path(d) / "spec.docx"
            real.write_bytes(b"placeholder")
            same1 = real
            same2 = (Path(d) / "sub" / ".." / "spec.docx")
            deduped = dedupe_paths([same1, same2])
            self.assertEqual(len(deduped), 1)

    def test_preserves_order(self) -> None:
        out = dedupe_paths([Path("/x/a"), Path("/x/b"), Path("/x/a"), Path("/x/c")])
        self.assertEqual(
            [p.name for p in out],
            ["a", "b", "c"],
        )

    def test_is_duplicate_of_any(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            real = Path(d) / "spec.docx"
            real.write_bytes(b"x")
            existing = [real]
            other = Path(d) / "sub" / ".." / "spec.docx"
            self.assertTrue(is_duplicate_of_any(other, existing))
            self.assertFalse(is_duplicate_of_any(Path(d) / "different.docx", existing))


class TestActorsTemplate(unittest.TestCase):
    def test_generates_readable_xlsx(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "nested" / "actors_template.xlsx"
            result = write_actors_template(out)
            self.assertEqual(result, out)
            self.assertTrue(out.exists())

            wb = load_workbook(out)
            self.assertIn("Actors", wb.sheetnames)
            self.assertIn("Readme", wb.sheetnames)

            ws = wb["Actors"]
            # Header row + 4 example rows.  Headers must match what
            # actors.load_actors_from_xlsx expects ("Actor" / "Aliases").
            self.assertEqual(ws.cell(row=1, column=1).value, "Actor")
            self.assertEqual(ws.cell(row=1, column=2).value, "Aliases")
            for i, (canonical, aliases) in enumerate(ACTORS_TEMPLATE_ROWS, start=2):
                self.assertEqual(ws.cell(row=i, column=1).value, canonical)
                self.assertEqual(ws.cell(row=i, column=2).value, aliases)

    def test_template_is_loadable_by_actor_loader(self) -> None:
        """Round-trip: generated template should parse via load_actors_from_xlsx."""
        from requirements_extractor.actors import load_actors_from_xlsx

        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "actors.xlsx"
            write_actors_template(out)
            actors = load_actors_from_xlsx(out)
            # Every canonical we seeded must come back.  ActorEntry uses
            # ``name`` as its canonical field, so match against that.
            names = {a.name for a in actors}
            for canonical, _ in ACTORS_TEMPLATE_ROWS:
                self.assertIn(canonical, names)


if __name__ == "__main__":
    unittest.main()
