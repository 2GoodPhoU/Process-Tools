"""Tests for the diff subcommand (REVIEW §3.12).

Covers:

1. ``compute_diff`` on dicts (pure function — easy to pin behaviour).
2. ``diff_workbooks`` end-to-end: write two small xlsx fixtures via
   the normal writer, run the diff, inspect the produced workbook.

Run:  python -m unittest tests.test_diff
"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from requirements_extractor.diff import (
    DiffEntry,
    _read_requirements_workbook,
    compute_diff,
    diff_workbooks,
    write_diff_workbook,
)
from requirements_extractor.models import Requirement, compute_stable_id
from requirements_extractor.writer import write_requirements


def _make_req(
    text: str,
    *,
    order: int = 1,
    source_file: str = "spec.docx",
    primary_actor: str = "User",
    row_ref: str = "Table 1, Row 3",
) -> Requirement:
    return Requirement(
        order=order,
        source_file=source_file,
        heading_trail="",
        section_topic="",
        row_ref=row_ref,
        block_ref="Paragraph 1",
        primary_actor=primary_actor,
        secondary_actors=[],
        text=text,
        req_type="Hard",
        keywords=["shall"],
        confidence="High",
        stable_id=compute_stable_id(source_file, primary_actor, text),
    )


def _row_dict(req: Requirement) -> dict:
    """Build the dict shape that compute_diff expects."""
    return {
        "ID": req.stable_id,
        "Source File": req.source_file,
        "Row Ref": req.row_ref,
        "Primary Actor": req.primary_actor,
        "Requirement": req.text,
    }


# ---------------------------------------------------------------------------
# compute_diff
# ---------------------------------------------------------------------------


class TestComputeDiffEmpty(unittest.TestCase):
    def test_empty_old_empty_new(self) -> None:
        self.assertEqual(compute_diff([], []), [])

    def test_empty_old_only_new(self) -> None:
        new = [_row_dict(_make_req("A"))]
        entries = compute_diff([], new)
        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0].change_type, "Added")

    def test_empty_new_only_old(self) -> None:
        old = [_row_dict(_make_req("A"))]
        entries = compute_diff(old, [])
        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0].change_type, "Removed")


class TestComputeDiffByStableId(unittest.TestCase):
    def test_identical_workbooks_produce_no_diff(self) -> None:
        rows = [_row_dict(_make_req("A")), _row_dict(_make_req("B", order=2))]
        self.assertEqual(compute_diff(rows, rows), [])

    def test_pure_add(self) -> None:
        a = _make_req("A")
        b = _make_req("B", order=2)
        old = [_row_dict(a)]
        new = [_row_dict(a), _row_dict(b)]
        entries = compute_diff(old, new)
        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0].change_type, "Added")
        self.assertEqual(entries[0].text, "B")

    def test_pure_remove(self) -> None:
        a = _make_req("A")
        b = _make_req("B", order=2)
        old = [_row_dict(a), _row_dict(b)]
        new = [_row_dict(a)]
        entries = compute_diff(old, new)
        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0].change_type, "Removed")
        self.assertEqual(entries[0].text, "B")


class TestComputeDiffChanged(unittest.TestCase):
    """Changed rows are detected by secondary (file, row_ref) match."""

    def test_text_edit_at_same_position_is_changed(self) -> None:
        old_req = _make_req("The User shall log in.")
        new_req = _make_req(
            "The User shall authenticate before logging in.",
            # Note: same file + same row_ref as old_req; different text.
            row_ref="Table 1, Row 3",
        )
        entries = compute_diff([_row_dict(old_req)], [_row_dict(new_req)])
        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0].change_type, "Changed")
        self.assertIn("log in", entries[0].old_text)
        self.assertIn("authenticate", entries[0].text)

    def test_text_edit_at_different_position_is_add_plus_remove(self) -> None:
        # Author moved the requirement to a new row ref AND edited the
        # text — we conservatively emit Add + Remove rather than a
        # false Changed pairing.
        old_req = _make_req("The User shall log in.", row_ref="Row 3")
        new_req = _make_req("The User shall authenticate.", row_ref="Row 9")
        entries = compute_diff([_row_dict(old_req)], [_row_dict(new_req)])
        types = sorted(e.change_type for e in entries)
        self.assertEqual(types, ["Added", "Removed"])


class TestComputeDiffOrdering(unittest.TestCase):
    """Output entries group Removed → Changed → Added for easy scanning."""

    def test_output_order_groups_change_types(self) -> None:
        # Distinct row_refs so the secondary (file, row_ref) match
        # doesn't cross-pair unrelated rows.
        a = _make_req("A", row_ref="Table 1, Row 1")
        b_old = _make_req("B", order=2, row_ref="Table 1, Row 2")
        b_new = _make_req("B prime", order=2, row_ref="Table 1, Row 2")
        c = _make_req("C", order=3, row_ref="Table 1, Row 3")
        d = _make_req("D", order=4, row_ref="Table 1, Row 4")

        old = [_row_dict(a), _row_dict(b_old), _row_dict(c)]
        new = [_row_dict(a), _row_dict(b_new), _row_dict(d)]
        entries = compute_diff(old, new)
        types = [e.change_type for e in entries]
        # Removed (C) first, Changed (B) next, Added (D) last.
        self.assertEqual(types, ["Removed", "Changed", "Added"])


# ---------------------------------------------------------------------------
# write_diff_workbook
# ---------------------------------------------------------------------------


class TestWriteDiffWorkbook(unittest.TestCase):
    def test_empty_entries_produces_header_only(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "diff.xlsx"
            write_diff_workbook([], out)
            self.assertTrue(out.exists())
            wb = load_workbook(out)
            ws = wb["Diff"]
            headers = [c.value for c in ws[1]]
            self.assertEqual(headers[0], "Change Type")
            # No body rows.
            self.assertEqual(ws.max_row, 1)

    def test_colour_per_change_type(self) -> None:
        entries = [
            DiffEntry("Added",   "REQ-aaaaaaaa", "a.docx", "R1", "U", "A text"),
            DiffEntry("Removed", "REQ-bbbbbbbb", "b.docx", "R1", "U", "B text"),
            DiffEntry("Changed", "REQ-cccccccc", "c.docx", "R1", "U",
                      "C new", old_text="C old"),
        ]
        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "diff.xlsx"
            write_diff_workbook(entries, out)
            wb = load_workbook(out)
            ws = wb["Diff"]
            # Row 2+ (after header) — the fills should differ.
            fills = [ws.cell(row=r, column=1).fill.start_color.rgb
                     for r in range(2, 5)]
            # Three distinct fills.
            self.assertEqual(len(set(fills)), 3)

    def test_summary_sheet_has_counts(self) -> None:
        entries = [
            DiffEntry("Added", "REQ-1", "a.docx", "R1", "U", "x"),
            DiffEntry("Added", "REQ-2", "a.docx", "R2", "U", "y"),
            DiffEntry("Removed", "REQ-3", "a.docx", "R3", "U", "z"),
        ]
        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "diff.xlsx"
            write_diff_workbook(entries, out)
            wb = load_workbook(out)
            self.assertIn("Summary", wb.sheetnames)
            summary = wb["Summary"]
            # Row order is Added, Removed, Changed.
            self.assertEqual(summary.cell(row=2, column=1).value, "Added")
            self.assertEqual(summary.cell(row=2, column=2).value, 2)
            self.assertEqual(summary.cell(row=3, column=1).value, "Removed")
            self.assertEqual(summary.cell(row=3, column=2).value, 1)
            self.assertEqual(summary.cell(row=4, column=1).value, "Changed")
            self.assertEqual(summary.cell(row=4, column=2).value, 0)


# ---------------------------------------------------------------------------
# diff_workbooks — full round trip through the regular writer
# ---------------------------------------------------------------------------


class TestDiffWorkbooksRoundTrip(unittest.TestCase):
    def test_end_to_end(self) -> None:
        # Build two requirement lists — old has A+B, new has A+B' (edit)+C.
        a = _make_req("The User shall authenticate.", order=1)
        b_old = _make_req("Response within 300 ms.", order=2,
                          row_ref="Table 2, Row 1")
        b_new = _make_req("Response within 150 ms.", order=2,
                          row_ref="Table 2, Row 1")
        c = _make_req("The Admin may revoke access.", order=3,
                      row_ref="Table 2, Row 2", primary_actor="Admin")

        with tempfile.TemporaryDirectory() as d:
            old_xlsx = Path(d) / "old.xlsx"
            new_xlsx = Path(d) / "new.xlsx"
            diff_xlsx = Path(d) / "diff.xlsx"
            write_requirements([a, b_old], old_xlsx)
            write_requirements([a, b_new, c], new_xlsx)
            out, counts = diff_workbooks(old_xlsx, new_xlsx, diff_xlsx)

            self.assertTrue(out.exists())
            self.assertEqual(counts, {"Added": 1, "Removed": 0, "Changed": 1})

            # Inspect the written workbook.
            wb = load_workbook(out)
            ws = wb["Diff"]
            # Header + 2 body rows.
            self.assertEqual(ws.max_row, 3)
            change_types = [
                ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)
            ]
            # Order: Removed -> Changed -> Added.
            self.assertEqual(change_types, ["Changed", "Added"])


class TestReadWorkbookSafety(unittest.TestCase):
    """Pin the two safety improvements in ``_read_requirements_workbook``:
    preferring the ``Requirements`` sheet by name, and raising a clear
    error when the workbook has no ``ID`` column (pre-stable-id format).
    """

    def test_raises_on_missing_id_column(self) -> None:
        from openpyxl import Workbook
        with tempfile.TemporaryDirectory() as d:
            path = Path(d) / "old.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Requirements"
            ws.append(["#", "Source File", "Requirement"])
            ws.append([1, "spec.docx", "The system shall log in."])
            wb.save(path)
            with self.assertRaises(ValueError) as ctx:
                _read_requirements_workbook(path)
            self.assertIn("ID", str(ctx.exception))
            self.assertIn("older than the stable-ID", str(ctx.exception))

    def test_prefers_requirements_sheet_over_active(self) -> None:
        """If the user saved the xlsx with the Summary tab active, the
        reader must still pick up the Requirements sheet."""
        from openpyxl import Workbook
        with tempfile.TemporaryDirectory() as d:
            path = Path(d) / "w.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Requirements"
            ws.append(["#", "ID", "Source File", "Requirement"])
            ws.append([1, "REQ-aaaaaaaa", "spec.docx", "The system shall X."])
            summary = wb.create_sheet("Summary")
            summary.append(["Total", 1])
            # Simulate "user left Summary active on save".
            wb.active = wb.sheetnames.index("Summary")
            wb.save(path)

            rows = _read_requirements_workbook(path)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["ID"], "REQ-aaaaaaaa")


if __name__ == "__main__":
    unittest.main()
