"""Write the coverage matrix to xlsx.

Output shape:

- **Sheet 1 — "Matrix"**: contract requirements as rows, procedure clauses
  as columns. Cell value is the rounded combined score (or empty when no
  matcher fired). Cell fill colour graduates from white (0) through
  yellow (medium) to green (high).
- **Sheet 2 — "Detail"**: one row per non-zero (contract, procedure)
  pair, with columns: contract_id, contract_text, procedure_id,
  procedure_text, score, matchers, evidence. This is the auditor's view
  — every link with the reasoning that voted for it.
- **Sheet 3 — "Gaps"**: requirements with zero matches and clauses with
  zero matches, side-by-side. The "what's missing" lens.

Frozen panes on Sheet 1 keep the requirement text and clause IDs visible
when scrolling.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import CombinedMatch, DDERow


_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
_GAP_FILL = PatternFill("solid", fgColor="FFE6E6")


def _score_fill(score: float) -> PatternFill | None:
    """Cell fill scaled by score. White → yellow → green."""

    if score <= 0:
        return None
    # Interpolate a hex colour from yellow (FFF59D) at 0.2 to green
    # (4CAF50) at 1.0. Below 0.2 leave plain.
    if score < 0.2:
        return PatternFill("solid", fgColor="FFF9C4")
    if score < 0.5:
        return PatternFill("solid", fgColor="FFF59D")
    if score < 0.75:
        return PatternFill("solid", fgColor="AED581")
    return PatternFill("solid", fgColor="4CAF50")


def write_matrix(
    contract_rows: List[DDERow],
    procedure_rows: List[DDERow],
    combined: Dict[Tuple[str, str], CombinedMatch],
    output_path: str | Path,
) -> None:
    output_path = Path(output_path)
    wb = Workbook()

    _write_matrix_sheet(wb.active, contract_rows, procedure_rows, combined)
    _write_detail_sheet(wb.create_sheet("Detail"), contract_rows, procedure_rows, combined)
    _write_gaps_sheet(wb.create_sheet("Gaps"), contract_rows, procedure_rows, combined)

    wb.save(output_path)


def _write_matrix_sheet(
    ws,
    contract_rows: List[DDERow],
    procedure_rows: List[DDERow],
    combined: Dict[Tuple[str, str], CombinedMatch],
) -> None:
    ws.title = "Matrix"

    # Header row: blank, blank, then procedure ids
    ws.cell(row=1, column=1, value="Contract ID").font = _HEADER_FONT
    ws.cell(row=1, column=2, value="Requirement").font = _HEADER_FONT
    ws.cell(row=1, column=1).fill = _HEADER_FILL
    ws.cell(row=1, column=2).fill = _HEADER_FILL

    for col_idx, clause in enumerate(procedure_rows, start=3):
        cell = ws.cell(row=1, column=col_idx, value=clause.stable_id)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(text_rotation=60, horizontal="center")

    # Body rows
    for row_idx, req in enumerate(contract_rows, start=2):
        ws.cell(row=row_idx, column=1, value=req.stable_id).font = _HEADER_FONT
        ws.cell(row=row_idx, column=2, value=req.text).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        for col_idx, clause in enumerate(procedure_rows, start=3):
            record = combined.get((req.stable_id, clause.stable_id))
            if record is None:
                continue
            cell = ws.cell(row=row_idx, column=col_idx, value=round(record.score, 2))
            fill = _score_fill(record.score)
            if fill is not None:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    # Sizing & freeze
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 60
    for col_idx in range(3, 3 + len(procedure_rows)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 8
    ws.row_dimensions[1].height = 90
    ws.freeze_panes = "C2"


def _write_detail_sheet(
    ws,
    contract_rows: List[DDERow],
    procedure_rows: List[DDERow],
    combined: Dict[Tuple[str, str], CombinedMatch],
) -> None:
    contract_by_id = {r.stable_id: r for r in contract_rows}
    procedure_by_id = {r.stable_id: r for r in procedure_rows}

    headers = [
        "Contract ID",
        "Requirement",
        "Procedure ID",
        "Clause",
        "Score",
        "Matchers",
        "Evidence",
    ]
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    sorted_pairs = sorted(
        combined.values(),
        key=lambda m: (-m.score, m.contract_id, m.procedure_id),
    )

    for row_idx, record in enumerate(sorted_pairs, start=2):
        contract = contract_by_id.get(record.contract_id)
        clause = procedure_by_id.get(record.procedure_id)
        ws.cell(row=row_idx, column=1, value=record.contract_id)
        ws.cell(row=row_idx, column=2, value=contract.text if contract else "")
        ws.cell(row=row_idx, column=3, value=record.procedure_id)
        ws.cell(row=row_idx, column=4, value=clause.text if clause else "")
        ws.cell(row=row_idx, column=5, value=round(record.score, 3))
        ws.cell(row=row_idx, column=6, value=", ".join(record.matchers))
        ws.cell(row=row_idx, column=7, value="\n".join(record.evidence))

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 60
    ws.freeze_panes = "A2"


def _write_gaps_sheet(
    ws,
    contract_rows: List[DDERow],
    procedure_rows: List[DDERow],
    combined: Dict[Tuple[str, str], CombinedMatch],
) -> None:
    matched_contract_ids = {key[0] for key in combined.keys()}
    matched_procedure_ids = {key[1] for key in combined.keys()}

    uncovered_reqs = [r for r in contract_rows if r.stable_id not in matched_contract_ids]
    unused_clauses = [r for r in procedure_rows if r.stable_id not in matched_procedure_ids]

    # Two side-by-side sections.
    ws.cell(row=1, column=1, value="Requirements with no procedure match").font = _HEADER_FONT
    ws.cell(row=1, column=4, value="Clauses with no requirement match").font = _HEADER_FONT
    ws.cell(row=1, column=1).fill = _GAP_FILL
    ws.cell(row=1, column=4).fill = _GAP_FILL

    sub_headers_left = ["Contract ID", "Source File", "Requirement"]
    sub_headers_right = ["Procedure ID", "Source File", "Clause"]
    for col_idx, name in enumerate(sub_headers_left, start=1):
        cell = ws.cell(row=2, column=col_idx, value=name)
        cell.font = _HEADER_FONT
    for col_idx, name in enumerate(sub_headers_right, start=4):
        cell = ws.cell(row=2, column=col_idx, value=name)
        cell.font = _HEADER_FONT

    for row_idx, req in enumerate(uncovered_reqs, start=3):
        ws.cell(row=row_idx, column=1, value=req.stable_id)
        ws.cell(row=row_idx, column=2, value=req.source_file or "")
        ws.cell(row=row_idx, column=3, value=req.text)
    for row_idx, clause in enumerate(unused_clauses, start=3):
        ws.cell(row=row_idx, column=4, value=clause.stable_id)
        ws.cell(row=row_idx, column=5, value=clause.source_file or "")
        ws.cell(row=row_idx, column=6, value=clause.text)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 60
    ws.freeze_panes = "A3"
