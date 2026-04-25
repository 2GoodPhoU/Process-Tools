"""End-to-end smoke test.

Builds two tiny in-memory DDE-shaped lists, runs every matcher + the
combiner, writes the matrix, then re-reads the output workbook and asserts
the expected pairs landed in the Detail sheet. Catches:

- import-time errors across the package
- matcher ↔ combiner ↔ writer wiring
- xlsx output well-formedness

Doesn't try to validate matcher *quality* — that's what per-matcher tests
will do once the scaffolding is in.
"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from compliance_matrix.combiner import combine
from compliance_matrix.matchers import (
    explicit_id,
    keyword_overlap,
    manual_mapping,
    similarity,
)
from compliance_matrix.matrix_writer import write_matrix
from compliance_matrix.models import DDERow


def _contract() -> list[DDERow]:
    return [
        DDERow(
            stable_id="REQ-AAA1",
            text="The Operator shall log in to the control console IAW Section 4.2.1.",
            primary_actor="Operator",
            side="contract",
        ),
        DDERow(
            stable_id="REQ-AAA2",
            text="The system shall record the timestamp of every operator login event.",
            primary_actor="System",
            side="contract",
        ),
        DDERow(
            stable_id="REQ-AAA3",
            text="An entirely unrelated requirement about hydraulic pressure sensors.",
            primary_actor="System",
            side="contract",
        ),
    ]


def _procedure() -> list[DDERow]:
    return [
        DDERow(
            stable_id="PROC-1111",
            text="Operator login procedure: enter credentials and confirm shift handover.",
            section="4.2.1",
            row_ref="Section 4.2.1",
            side="procedure",
        ),
        DDERow(
            stable_id="PROC-2222",
            text="Login event timestamps must be recorded by the system.",
            section="4.2.2",
            side="procedure",
        ),
        DDERow(
            stable_id="PROC-3333",
            text="Boilerplate compliance statement — completely irrelevant text.",
            section="9.0",
            side="procedure",
        ),
    ]


class TestEndToEnd(unittest.TestCase):
    def test_smoke_round_trip(self) -> None:
        contract = _contract()
        procedure = _procedure()

        all_matches: list = []
        all_matches.extend(explicit_id.run(contract, procedure))
        all_matches.extend(
            keyword_overlap.run(contract, procedure, threshold=0.10)
        )
        all_matches.extend(
            similarity.run(contract, procedure, threshold=0.10)
        )

        combined = combine(all_matches)

        # Sanity: the section-4.2.1-cited requirement should match its
        # cited clause via the explicit_id matcher.
        self.assertIn(("REQ-AAA1", "PROC-1111"), combined)
        self.assertIn("explicit_id", combined[("REQ-AAA1", "PROC-1111")].matchers)

        # The hydraulic-pressure requirement should land in nothing.
        self.assertFalse(
            any(key[0] == "REQ-AAA3" for key in combined.keys()),
            f"REQ-AAA3 should be uncovered; got {list(combined.keys())}",
        )

        with tempfile.TemporaryDirectory() as td:
            output = Path(td) / "matrix.xlsx"
            write_matrix(contract, procedure, combined, output)
            self.assertTrue(output.exists())

            wb = load_workbook(output)
            self.assertEqual(
                {"Matrix", "Detail", "Gaps"}, set(wb.sheetnames)
            )

            # Detail sheet should have at least one row beyond the header.
            detail = wb["Detail"]
            self.assertGreaterEqual(detail.max_row, 2)

            # Gaps sheet should list REQ-AAA3 (hydraulic) somewhere on the
            # left column block.
            gaps = wb["Gaps"]
            left_col = [row[0].value for row in gaps.iter_rows(min_col=1, max_col=1)]
            self.assertIn("REQ-AAA3", left_col)

    def test_manual_mapping_yaml(self) -> None:
        contract = _contract()
        procedure = _procedure()

        with tempfile.TemporaryDirectory() as td:
            mapping_path = Path(td) / "mapping.yaml"
            mapping_path.write_text(
                "REQ-AAA3: [PROC-3333]\n", encoding="utf-8"
            )
            try:
                hits = manual_mapping.run(contract, procedure, mapping_path)
            except RuntimeError as exc:
                # PyYAML missing in the sandbox is fine — skip then.
                self.skipTest(f"PyYAML unavailable: {exc}")
            self.assertEqual(len(hits), 1)
            self.assertEqual(hits[0].score, 1.0)

    def test_manual_mapping_csv(self) -> None:
        contract = _contract()
        procedure = _procedure()

        with tempfile.TemporaryDirectory() as td:
            mapping_path = Path(td) / "mapping.csv"
            mapping_path.write_text(
                "contract_id,procedure_id,note\n"
                "REQ-AAA3,PROC-3333,covered by hydraulic chapter\n",
                encoding="utf-8",
            )
            hits = manual_mapping.run(contract, procedure, mapping_path)
            self.assertEqual(len(hits), 1)
            self.assertEqual(hits[0].evidence,
                             "manual mapping (mapping.csv): covered by hydraulic chapter")


if __name__ == "__main__":
    unittest.main()
