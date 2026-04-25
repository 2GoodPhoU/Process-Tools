"""Side-car xlsx listing every flagged activity for human judgment.

The skeleton diagram is best-effort. Whenever the builder wasn't sure
about a node — negative-polarity requirements, conditional clauses
without obvious branch targets, requirements with no modal verb —
it tags ``flagged=True`` on the activity. This module turns those
flags into a single-sheet review workbook.

Columns: Stable ID, Actor, Label (as rendered in the diagram), Flag
reason, original requirement text (if available), and a Reviewer Decision
column (left empty for the human to fill in: 'keep', 'drop', 'rephrase',
etc.).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .models import Skeleton


_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
_FLAG_FILL = PatternFill("solid", fgColor="FFE6E6")
_ITALIC_FONT = Font(italic=True)


def write_review(skeleton: Skeleton, output_path, dde_rows=None) -> None:
    output_path = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"

    headers = [
        "Stable ID",
        "Actor",
        "Label",
        "Flag Reason",
        "Source Requirement",
        "Reviewer Decision",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    flagged = skeleton.review_records()
    
    # Build stable_id → requirement text lookup
    dde_lookup = {}
    if dde_rows:
        for row in dde_rows:
            if hasattr(row, 'stable_id') and hasattr(row, 'text'):
                dde_lookup[row.stable_id] = row.text
    if not flagged:
        ws.cell(row=2, column=1, value="(no flagged items — skeleton built cleanly)")
        ws.cell(row=2, column=1).font = _ITALIC_FONT
    else:
        for row_idx, activity in enumerate(flagged, start=2):
            ws.cell(row=row_idx, column=1, value=activity.stable_id)
            ws.cell(row=row_idx, column=2, value=activity.actor)
            ws.cell(row=row_idx, column=3, value=activity.label)
            ws.cell(row=row_idx, column=4, value=activity.flag_reason or "")
            ws.cell(row=row_idx, column=5, value=dde_lookup.get(activity.stable_id, ""))
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).fill = _FLAG_FILL

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 70  # Source requirement
    ws.column_dimensions["F"].width = 28  # Reviewer decision
    ws.freeze_panes = "A2"

    wb.save(output_path)
