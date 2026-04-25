"""Unit tests for the enhanced review writer."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from nimbus_skeleton.models import Activity, Skeleton, DDERow
from nimbus_skeleton.review_writer import write_review


class TestReviewWriterEnhanced(unittest.TestCase):
    """Test review sheet generation with source requirement text."""

    def test_review_sheet_with_source_text(self):
        """Review sheet includes original requirement text."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "review.xlsx"

            # Build a skeleton with flagged activities
            skeleton = Skeleton()
            skeleton.add_activity(Activity(
                stable_id="REQ-001",
                label="Operator logs in",
                actor="Operator",
                flagged=True,
                flag_reason="negative polarity",
            ))

            # Mock DDE rows with requirement text
            dde_rows = [
                DDERow(
                    stable_id="REQ-001",
                    text="The operator shall NOT bypass the safety check.",
                    primary_actor="Operator",
                ),
            ]

            # Write review with source text
            write_review(skeleton, output_path, dde_rows=dde_rows)

            # Verify file was created and contains the text
            assert output_path.exists(), "Review xlsx not created"

            wb = load_workbook(output_path)
            ws = wb["Review"]

            # Check header row includes source requirement column
            headers = [cell.value for cell in ws[1]]
            assert "Source Requirement" in headers, f"Missing source column. Headers: {headers}"

            # Check that the requirement text was written
            source_col_idx = headers.index("Source Requirement") + 1
            req_text = ws.cell(row=2, column=source_col_idx).value
            assert req_text == "The operator shall NOT bypass the safety check.", \
                f"Source text not found: {req_text}"

    def test_review_sheet_without_dde_rows(self):
        """Review sheet works fine when dde_rows is None."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "review.xlsx"

            skeleton = Skeleton()
            skeleton.add_activity(Activity(
                stable_id="REQ-002",
                label="System records timestamp",
                actor="System",
                flagged=True,
                flag_reason="no modal verb",
            ))

            # Write review WITHOUT source rows (backward compat)
            write_review(skeleton, output_path, dde_rows=None)

            assert output_path.exists()
            wb = load_workbook(output_path)
            ws = wb["Review"]

            # The cell should just be empty (not error)
            headers = [cell.value for cell in ws[1]]
            source_col_idx = headers.index("Source Requirement") + 1
            source_text = ws.cell(row=2, column=source_col_idx).value
            assert source_text is None or source_text == "", \
                f"Expected empty source, got: {source_text}"

    def test_review_sheet_no_flagged_items(self):
        """Review sheet shows empty message when no flagged items."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "review.xlsx"

            skeleton = Skeleton()
            # No activities added, so nothing is flagged

            write_review(skeleton, output_path, dde_rows=[])

            assert output_path.exists()
            wb = load_workbook(output_path)
            ws = wb["Review"]

            # Should have header + one message row
            assert ws.cell(row=2, column=1).value is not None, "No message row for empty review"


if __name__ == "__main__":
    unittest.main()
