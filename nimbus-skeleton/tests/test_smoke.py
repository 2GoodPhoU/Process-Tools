"""End-to-end smoke test for the Nimbus Skeleton Mapper.

Builds a small DDE-row corpus exercising every classifier branch (plain
imperative, conditional, declarative, negative-polarity), runs the
builder, renders both emitters and the review side-car, and asserts on
the structure of each output. Catches:

- import-time errors across the package
- builder ↔ emitter wiring
- classifier coverage on the basic cases
- xlsx + plantuml + manifest output well-formedness
"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from nimbus_skeleton.builder import build_skeleton
from nimbus_skeleton.classifier import classify
from nimbus_skeleton.emitters import manifest, plantuml
from nimbus_skeleton.models import DDERow
from nimbus_skeleton.review_writer import write_review


def _corpus():
    return [
        DDERow(
            stable_id="REQ-001",
            text="The Operator shall log in to the control console.",
            primary_actor="Operator",
            polarity="Positive",
        ),
        DDERow(
            stable_id="REQ-002",
            text="The System shall record a timestamp for every login event.",
            primary_actor="System",
            polarity="Positive",
        ),
        DDERow(
            stable_id="REQ-003",
            text="If the login fails, the System shall lock the account.",
            primary_actor="System",
            polarity="Positive",
        ),
        DDERow(
            stable_id="REQ-004",
            text="The Operator shall not bypass the safety interlock.",
            primary_actor="Operator",
            polarity="Negative",
        ),
        DDERow(
            stable_id="REQ-005",
            text="The audit log is defined as the persistent record of all "
            "system events.",
            primary_actor="System",
            polarity="Positive",
        ),
    ]


class TestClassifier(unittest.TestCase):
    def test_imperative(self) -> None:
        cls = classify("The Operator shall verify the indicators.")
        self.assertEqual(cls.kind, "activity")
        self.assertFalse(cls.flagged)

    def test_conditional(self) -> None:
        cls = classify("If verification fails, the System shall halt.")
        self.assertEqual(cls.kind, "gateway")

    def test_declarative(self) -> None:
        cls = classify("The audit log is defined as the persistent record.")
        self.assertEqual(cls.kind, "note")

    def test_negative_polarity_flagged(self) -> None:
        cls = classify("Operator shall not override the lock.", polarity="Negative")
        self.assertEqual(cls.kind, "activity")
        self.assertTrue(cls.flagged)


class TestBuilder(unittest.TestCase):
    def test_actors_and_swimlane_order(self) -> None:
        skeleton = build_skeleton(_corpus())
        # Operator + System should both be actors; the System gateway
        # node should be in skeleton.gateways.
        self.assertIn("Operator", skeleton.actors)
        self.assertIn("System", skeleton.actors)
        self.assertEqual(len(skeleton.gateways), 1)
        self.assertEqual(skeleton.gateways[0].actor, "System")

    def test_negative_polarity_is_flagged(self) -> None:
        skeleton = build_skeleton(_corpus())
        bypass = next(a for a in skeleton.activities if a.stable_id == "REQ-004")
        self.assertTrue(bypass.flagged)

    def test_declarative_becomes_note(self) -> None:
        skeleton = build_skeleton(_corpus())
        self.assertTrue(any(n.stable_id == "REQ-005" for n in skeleton.notes))

    def test_alias_resolution(self) -> None:
        rows = [
            DDERow(stable_id="R1", text="The op shall press start.",
                   primary_actor="op", polarity="Positive"),
            DDERow(stable_id="R2", text="The Operator shall confirm.",
                   primary_actor="Operator", polarity="Positive"),
        ]
        skeleton = build_skeleton(
            rows,
            actors_overrides={"Operator": ["op", "the operator"]},
        )
        # Both rows should fold into a single canonical Operator swimlane.
        self.assertEqual(skeleton.actors, ["Operator"])


class TestEmitters(unittest.TestCase):
    def test_plantuml_renders(self) -> None:
        skeleton = build_skeleton(_corpus())
        out = plantuml.render(skeleton)
        self.assertIn("@startuml", out)
        self.assertIn("@enduml", out)
        self.assertIn("|Operator|", out)
        self.assertIn("|System|", out)
        # Gateway should appear as if/endif.
        self.assertIn("if (", out)
        self.assertIn("endif", out)

    def test_manifest_renders(self) -> None:
        skeleton = build_skeleton(_corpus())
        out = manifest.render(skeleton)
        # Whether yaml or json fallback, the content keys must appear.
        for key in ("actors", "activities", "gateways", "flows", "notes"):
            self.assertIn(key, out)

    def test_review_writer(self) -> None:
        skeleton = build_skeleton(_corpus())
        with tempfile.TemporaryDirectory() as td:
            output = Path(td) / "review.xlsx"
            write_review(skeleton, output)
            self.assertTrue(output.exists())
            wb = load_workbook(output)
            ws = wb.active
            # Header + at least one flagged row (REQ-004 negative-polarity).
            self.assertEqual(ws.cell(row=1, column=1).value, "Stable ID")
            self.assertGreaterEqual(ws.max_row, 2)




class TestXmiEmitter(unittest.TestCase):
    def test_xmi_renders_valid_uml(self) -> None:
        from xml.etree import ElementTree as ET

        from nimbus_skeleton.emitters import xmi

        skeleton = build_skeleton(_corpus())
        out = xmi.render(skeleton, title="Smoke Test Model")

        # Must parse as valid XML.
        root = ET.fromstring(out)
        self.assertTrue(root.tag.endswith("XMI"))

        # uml:Activity must contain at least the actions and partitions
        # we expect.
        ns = {"uml": "http://www.omg.org/spec/UML/20131001",
              "xmi": "http://www.omg.org/spec/XMI/20131001"}

        activities = []
        partitions = []
        for elem in root.iter():
            xmi_type = elem.get("{http://www.omg.org/spec/XMI/20131001}type", "")
            if xmi_type == "uml:OpaqueAction":
                activities.append(elem.get("name"))
            elif xmi_type == "uml:ActivityPartition":
                partitions.append(elem.get("name"))

        self.assertGreaterEqual(len(activities), 1)
        self.assertIn("Operator", partitions)
        self.assertIn("System", partitions)

    def test_xmi_byte_stable_across_runs(self) -> None:
        from nimbus_skeleton.emitters import xmi

        skeleton = build_skeleton(_corpus())
        first = xmi.render(skeleton, title="Stability Check")
        second = xmi.render(skeleton, title="Stability Check")
        self.assertEqual(first, second)



class TestVsdxEmitter(unittest.TestCase):
    def test_vsdx_is_valid_zip_with_required_parts(self) -> None:
        import zipfile

        from nimbus_skeleton.emitters import vsdx

        skeleton = build_skeleton(_corpus())
        with tempfile.TemporaryDirectory() as td:
            output = Path(td) / "out.vsdx"
            vsdx.write(skeleton, output, title="VSDX Smoke Test")
            self.assertTrue(output.exists())

            self.assertTrue(zipfile.is_zipfile(output))
            with zipfile.ZipFile(output) as zf:
                names = set(zf.namelist())
                # Required OOXML / Visio parts.
                for required in (
                    "[Content_Types].xml",
                    "_rels/.rels",
                    "docProps/app.xml",
                    "docProps/core.xml",
                    "visio/document.xml",
                    "visio/_rels/document.xml.rels",
                    "visio/pages/pages.xml",
                    "visio/pages/_rels/pages.xml.rels",
                    "visio/pages/page1.xml",
                ):
                    self.assertIn(required, names, f"missing part: {required}")

    def test_vsdx_uses_nimbus_compatible_shape_names(self) -> None:
        """Nimbus's Visio-import rules key off the NameU attribute."""

        import zipfile

        from nimbus_skeleton.emitters import vsdx

        skeleton = build_skeleton(_corpus())
        with tempfile.TemporaryDirectory() as td:
            output = Path(td) / "out.vsdx"
            vsdx.write(skeleton, output)
            with zipfile.ZipFile(output) as zf:
                page = zf.read("visio/pages/page1.xml").decode("utf-8")

        # Activities → "Process"; gateways → "Decision".
        self.assertIn('NameU="Process"', page)
        # The corpus has an `if`-clause, so we should see at least one
        # Decision shape.
        self.assertIn('NameU="Decision"', page)

    def test_vsdx_byte_stable_across_runs(self) -> None:
        from nimbus_skeleton.emitters import vsdx

        skeleton = build_skeleton(_corpus())
        with tempfile.TemporaryDirectory() as td:
            a = Path(td) / "a.vsdx"
            b = Path(td) / "b.vsdx"
            vsdx.write(skeleton, a, title="Stability")
            vsdx.write(skeleton, b, title="Stability")
            # Inner page1.xml should be identical even if outer zip
            # timestamps differ.
            import zipfile
            with zipfile.ZipFile(a) as zfa, zipfile.ZipFile(b) as zfb:
                self.assertEqual(
                    zfa.read("visio/pages/page1.xml"),
                    zfb.read("visio/pages/page1.xml"),
                )

if __name__ == "__main__":
    unittest.main()
